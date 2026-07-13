# Copyright (c) 2026, QTPL and contributors
# For license information, please see license.txt
from openpyxl.cell import read_only
import frappe
from frappe.model.document import Document
from frappe.utils import flt
from frappe.model.mapper import get_mapped_doc
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from collections import defaultdict
import re
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font
import math
import json

STEEL_BAR_DIAMETERS = [8, 10, 12, 16, 20, 25, 28, 32]
STEEL_BAR_WEIGHT_PER_METER = {d: round((d ** 2) / 162.0, 3) for d in STEEL_BAR_DIAMETERS}


def _get_row_reinforcement_weight(row):
    """Total weight (Kg) of one ra_steel_details row, summed across all populated diameter columns."""
    total = 0.0
    for d in STEEL_BAR_DIAMETERS:
        val = flt(row.get(f"{d}_mm_reinforcement"))
        if val:
            total += val * STEEL_BAR_WEIGHT_PER_METER[d]
    return total


def _get_deepest_task_id(row):
    """
    Returns the deepest populated task/level id on a row — used to match
    a ra_steel_details row to its corresponding ra_billing_details row.
    """
    level_fields = [f"task_level{i}" for i in range(1, 11)]
    for fieldname in reversed(level_fields):
        if row.get(fieldname):
            return row.get(fieldname)
    return row.get("task")

class RABilling(Document):

    def before_save(self):
        self.sync_deleted_tasks()
        self.sync_steel_quantities_to_billing()
        self.update_abstract_details()
        # steel_map = {}
        # for row in self.ra_steel_details:
        #     if row.subtask not in steel_map:
        #         steel_map[row.subtask] = 0
        #     steel_map[row.subtask] += (row.total_weight or 0)

        # for row in self.ra_billing_details:
        #     if row.subtask in steel_map:
        #         row.quantity = steel_map[row.subtask]
    def sync_steel_quantities_to_billing(self):
        """
        For every subtask that has reinforcement entries in ra_steel_details,
        compute its total weight (Kg), convert to Metric Tonne, and write
        that value into the matching row's quantity in ra_billing_details
        (matched by identical deepest task/level hierarchy).
        """
        steel_weight_kg_by_task = {}

        for row in self.ra_steel_details:
            deepest_id = _get_deepest_task_id(row)
            if not deepest_id:
                continue

            weight = _get_row_reinforcement_weight(row)
            if not weight:
                continue

            steel_weight_kg_by_task[deepest_id] = (
                steel_weight_kg_by_task.get(deepest_id, 0) + weight
            )

        if not steel_weight_kg_by_task:
            return

        for row in self.ra_billing_details:
            deepest_id = _get_deepest_task_id(row)

            if deepest_id in steel_weight_kg_by_task:
                weight_kg = steel_weight_kg_by_task[deepest_id]
                weight_mt = weight_kg / 1000.0

                row.quantity = weight_mt
                row.uom = "Metric Tonne"
                row.amount = flt(row.rate) * weight_mt
    def sync_deleted_tasks(self):
        if not self.is_new() and self.get_doc_before_save():
            old_tasks = set(d.task for d in self.get_doc_before_save().ra_abstract_details if d.task)
            current_tasks = set(d.task for d in self.ra_abstract_details if d.task)
            deleted_tasks = old_tasks - current_tasks
            
            if deleted_tasks:
                valid_details = [d for d in self.ra_billing_details if d.task not in deleted_tasks]
                self.set("ra_billing_details", valid_details)

    def on_submit(self):
        self.update_billed_quantity()
            
    def update_abstract_details(self):
            self.set("ra_abstract_details", [])

            stage_order = []
            stage_data = {}

            for row in self.ra_billing_details:
                key = row.stage

                if key not in stage_data:
                    stage_data[key] = {
                        "stage_subject": row.stage_subject,
                        "billed_quantity": 0.0,
                        "amount": 0.0,
                        "uom": row.uom,
                        "rate": flt(row.rate),
                        "description": "",
                    }
                    stage_order.append(key)
                
                stage_data[key]["billed_quantity"] += flt(row.quantity)
                stage_data[key]["amount"] += flt(row.amount)
                if row.uom:
                    stage_data[key]["uom"] = row.uom
                if flt(row.rate):
                    stage_data[key]["rate"] = flt(row.rate)
                if not stage_data[key]["description"]:
                    stage_data[key]["description"] = getattr(row, "description", "")

                # stage_data[key]["billed_quantity"] += flt(row.quantity)
                # stage_data[key]["amount"] += flt(row.amount)
                # stage_data[key]["uom"] = row.uom
                # stage_data[key]["rate"] = flt(row.rate)
                # if not stage_data[key]["description"]:
                #     stage_data[key]["description"] = getattr(row, "description", "")

            previous_totals = get_previous_stage_totals(self.project, self.name)

            for key in stage_order:
                data = stage_data[key]
                prev = previous_totals.get(key, {})
                previous_qty = flt(prev.get("total_bill_quantity"))
                previous_amt = flt(prev.get("total_bill_amount"))

                this_qty = data["billed_quantity"]
                this_amt = data["amount"]

                self.append("ra_abstract_details", {
                    "stage": key,
                    "stage_subject": data["stage_subject"],
                    "description": data["description"],
                    "uom": data["uom"],
                    "rate": data["rate"],
                    "previous_bill_quantity": previous_qty,
                    "billed_quantity": this_qty,
                    "total_bill_quantity": previous_qty + this_qty,
                    "previous_bill_amount": previous_amt,
                    "amount": this_amt,
                    "total_bill_amount": previous_amt + this_amt,
                })

            self.grand_total = sum(flt(d.amount) for d in self.ra_billing_details)

    # def update_abstract_details(self):
    #     self.set("ra_abstract_details", [])
        
    #     abstract_data = {}
    #     for row in self.ra_billing_details:
    #         key = (row.stage, row.stage_subject, row.task, row.task_subject, row.uom)
    #         if key not in abstract_data:
    #             abstract_data[key] = {
    #                 "billed_quantity": 0.0,
    #                 "rate": 0.0,
    #                 "amount": 0.0,
    #                 "description": getattr(row, "description", "")
    #             }
            
    #         abstract_data[key]["billed_quantity"] += flt(row.quantity)
    #         abstract_data[key]["rate"] = flt(row.rate)
    #         abstract_data[key]["amount"] += flt(row.amount)
    #         if not abstract_data[key]["description"]:
    #             abstract_data[key]["description"] = getattr(row, "description", "")
            
    #     for key, data in abstract_data.items():
    #         self.append("ra_abstract_details", {
    #             "stage": key[0],
    #             "stage_subject": key[1],
    #             "task": key[2],
    #             "task_subject": key[3],
    #             "uom": key[4],
    #             "rate": data["rate"],
    #             "billed_quantity": data["billed_quantity"],
    #             "amount": data["amount"],
    #             "description": data["description"]
    #         })
            
    #     self.grand_total = sum(flt(d.amount) for d in self.ra_billing_details)
    def update_billed_quantity(self):

        level_fields = [f"task_level{i}" for i in range(1, 11)]

        for row in self.ra_billing_details:

            # Find the deepest populated level - this is the "subtask" equivalent
            deepest_task = None
            for fieldname in reversed(level_fields):
                if row.get(fieldname):
                    deepest_task = row.get(fieldname)
                    break

            if not deepest_task:
                # fall back to task if no levels are set at all
                deepest_task = row.task

            if not deepest_task:
                continue

            current_billed = flt(
                frappe.db.get_value(
                    "Task",
                    deepest_task,
                    "custom_billed_quantity"
                )
            )

            frappe.db.set_value(
                "Task",
                deepest_task,
                "custom_billed_quantity",
                current_billed + flt(row.quantity)
            )

    # def update_billed_quantity(self):

    #     for row in self.ra_billing_details:

    #         if not row.subtask_subject:
    #             continue

    #         subtask_name = frappe.db.get_value(
    #             "Task",
    #             {
    #                 "subject": row.subtask_subject,
    #                 ""
    #                 "custom_is_subtask": 1,
    #                 "project": self.project
    #             },
    #             "name"
    #         )

    #         if not subtask_name:
    #             continue

    #         current_billed = flt(
    #             frappe.db.get_value(
    #                 "Task",
    #                 subtask_name,
    #                 "custom_billed_quantity"
    #             )
    #         )

    #         frappe.db.set_value(
    #             "Task",
    #             subtask_name,
    #             "custom_billed_quantity",
    #             current_billed + flt(row.quantity)
    #         )

@frappe.whitelist()
def validate_task_rates(doc):
    task_rate_map = {}
    doc = frappe.get_doc(frappe.parse_json(doc))
    for row in doc.ra_billing_details:
        task = (row.task or "").strip()
        rate = flt(row.rate)
        task_subject = frappe.db.get_value("Task",task, "subject")
        if task in task_rate_map:
            if task_rate_map[task] != rate:
                frappe.throw(
                    f"Task <b>{task_subject}</b> (Row {row.idx}) must have the same "
                    f"Rate: {task_rate_map[task]}, Current Rate: {rate}."
                )
        else:
            task_rate_map[task] = rate

@frappe.whitelist()
def get_project_tasks(project):
    """
    Fetch all subtasks for the selected project and return enriched data for RA Billing.
    """
    if not project:
        return []

    subtasks = frappe.get_all(
        "Task",
        filters={
            "project": project,
            "custom_is_subtask": 1
        },
        fields=[
            "name", "subject", "parent_task",
            "custom_total_quantity", "custom_total_achieved",
            "custom_rate", "custom_billed_quantity", "custom_uom"
        ]
    )

    if not subtasks:
        return []

    result = []
    for subtask in subtasks:
        task_id = subtask.parent_task
        task = None
        stage = None

        hierarchy = []

        current_task = subtask

        while current_task:

            hierarchy.insert(0, {
                "name": current_task.name,
                "subject": current_task.subject
            })

            if not current_task.parent_task:
                break

            current_task = frappe.db.get_value(
                "Task",
                current_task.parent_task,
                ["name", "subject", "parent_task"],
                as_dict=True
            )

        billed_qty = flt(subtask.custom_billed_quantity)
        rate = flt(subtask.custom_rate)
        achieved_qty = flt(subtask.custom_total_achieved)
        billable_qty = achieved_qty - billed_qty
        row = {
            "total_quantity": subtask.custom_total_quantity,
            "total_achieved": achieved_qty,
            "billed_quantity": billed_qty,
            "billable_quantity": billable_qty,
            "rate": flt(subtask.custom_rate),
            "uom": subtask.custom_uom
        }

        # Map hierarchy sequentially
        field_order = [
            "stage",
            "task",
            # "subtask",
            "task_level1",
            "task_level2",
            "task_level3",
            "task_level4",
            "task_level5",
            "task_level6",
            "task_level7",
            "task_level8",
            "task_level9",
            "task_level10"
        ]

        for idx, node in enumerate(hierarchy):

            if idx >= len(field_order):
                break

            fieldname = field_order[idx]

            row[fieldname] = node["subject"]
            row[f"{fieldname}_id"] = node["name"]
        result.append(row)
    return result

@frappe.whitelist()
def get_project_steel_tasks(project):
    if not project:
        return []

    subtasks = frappe.get_all(
        "Task",
        filters={
            "project": project,
            "custom_is_steel_subtask": 1
        },
        fields=[
            "name",
            "subject",
            "parent_task",
        ]
    )

    if not subtasks:
        return []

    result = []

    field_order = [
        "stage",
        "task",
        "task_level1",
        "task_level2",
        "task_level3",
        "task_level4",
        "task_level5",
        "task_level6",
        "task_level7",
        "task_level8",
        "task_level9",
        "task_level10"
    ]

    for subtask in subtasks:

        hierarchy = []
        current_task = subtask

        while current_task:

            hierarchy.insert(0, {
                "name": current_task.name,
                "subject": current_task.subject
            })

            if not current_task.parent_task:
                break

            current_task = frappe.db.get_value(
                "Task",
                current_task.parent_task,
                ["name", "subject", "parent_task"],
                as_dict=True
            )

        row = {}    
        for idx, node in enumerate(hierarchy):
            if idx >= len(field_order):
                break

            fieldname = field_order[idx]
            row[fieldname] = node["subject"]
            row[f"{fieldname}_id"] = node["name"]

        result.append(row)

    return result
    
@frappe.whitelist()
def create_sales_invoice(source_name, target_doc=None, item_code=None):
    if not item_code:
        args = getattr(frappe.flags, "args", None) or frappe.form_dict
        item_code = args.get("item_code")

    item_name = frappe.db.get_value("Item", item_code, "item_name")
    uom = frappe.db.get_value(
        "UOM Conversion Detail",
        {"parent": item_code},
        "uom"
    )
    if not uom and item_code:
        uom = frappe.db.get_value("Item", item_code, "stock_uom")
    
    def set_missing_values(source, target):
        target.customer = source.customer
        target.project = source.project
        target.custom_doc_link_doctype = "RA Billing"
        target.custom_doc_link = source.name
        
        target.append("items", {
            "item_code": item_code,
            "item_name": item_name,
            "qty": 1,
            "uom": uom,
            "rate": flt(source.grand_total),
            "price_list_rate": flt(source.grand_total),
            "amount": flt(source.grand_total),
            "base_rate": flt(source.grand_total),
            "base_amount": flt(source.grand_total),
            "base_price_list_rate": flt(source.grand_total)
        })
        
        target.run_method("set_missing_values")
       

    doc = get_mapped_doc(
        "RA Billing",
        source_name,
        {
            "RA Billing": {
                "doctype": "Sales Invoice"
            }
        },
        target_doc,
        set_missing_values
    )

    return doc       


@frappe.whitelist()
def export_ra_excel(ra_billing):
    doc = frappe.get_doc("RA Billing", ra_billing)
    project_doc = frappe.get_doc("Project", doc.project)
    company_name = project_doc.company

    wb = openpyxl.Workbook()
    wb.remove(wb.active)                     
    used_sheet_names = set()

    summary_ws = build_summary_sheet(wb, doc, used_sheet_names)

    # Sheet 1: Abstract
    ws_abstract = wb.create_sheet(_safe_sheet_name("Abstract", used_sheet_names))   

    # Sheet 2: Measurement
    ws_meas = wb.create_sheet(_safe_sheet_name("Measurement", used_sheet_names))
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Helper for caching descriptions
    task_desc_cache = {}
    steel_task_totals = {}
    steel_stage_totals = {}
    steel_stage_weights = {}   
    def get_task_desc(task_id):
        if not task_id: return ""
        if task_id not in task_desc_cache:
            desc = frappe.db.get_value("Task", task_id, "description")
            task_desc_cache[task_id] = frappe.utils.strip_html(desc) if desc else ""
        return task_desc_cache[task_id]

    # --- Weight-per-meter constants (kg/m), standard TMT formula: dia² / 162 ---
    BAR_DIAMETERS = [8, 10, 12, 16, 20, 25, 28, 32]
    BAR_WEIGHT_PER_METER = {d: round((d ** 2) / 162.0, 3) for d in BAR_DIAMETERS}

    def get_row_reinforcement(row):
        """Returns {dia: length_in_m} for this row's populated diameters."""
        values = {}
        for d in BAR_DIAMETERS:
            val = flt(row.get(f"{d}_mm_reinforcement"))
            if val:
                values[d] = val
        return values

    def get_row_weight(row):
        """Total weight (kg) for one steel row, summed across all diameters."""
        total = 0.0
        for d, length in get_row_reinforcement(row).items():
            total += length * BAR_WEIGHT_PER_METER[d]
        return total

    LEVEL_FIELDS = [f"task_level{i}" for i in range(1, 10)]  # level1..level9
    LEVEL_SUBJECT_FIELDS = [f"level{i}_subject" for i in range(1, 10)]

    def get_row_hierarchy(row):
        """
        Returns ordered list of (id, subject) tuples: stage, task, then
        whichever task_level1..9 are actually populated on this row.
        """
        path = []

        if row.get("stage"):
            path.append((row.get("stage"), row.get("stage_subject") or row.get("stage")))

        if row.get("task"):
            path.append((row.get("task"), row.get("task_subject") or row.get("task")))

        for lvl_field, subj_field in zip(LEVEL_FIELDS, LEVEL_SUBJECT_FIELDS):
            if row.get(lvl_field):
                path.append((row.get(lvl_field), row.get(subj_field) or row.get(lvl_field)))

        return path


    ws_steel = wb.create_sheet("Steel")

    steel_total_cols = 20  # Location, Nos, L, W, D, Ftg Depth, Dia, Spacing, Bar Nos, Length, 8x reinforcement cols, Remark
    last_steel_col = get_column_letter(steel_total_cols)

    ws_steel.merge_cells(f'A1:{last_steel_col}1')
    ws_steel['A1'] = company_name
    ws_steel['A1'].font = Font(bold=True, size=14)
    ws_steel['A1'].alignment = center_align

    ws_steel.merge_cells(f'A2:{last_steel_col}2')
    ws_steel['A2'] = project_doc.project_name or project_doc.name
    ws_steel['A2'].font = Font(bold=True, size=12)
    ws_steel['A2'].alignment = center_align

    ws_steel.merge_cells(f'A3:{last_steel_col}3')
    ws_steel['A3'] = f"RA Bill Number: {doc.name}"
    ws_steel['A3'].font = Font(bold=True, size=12)
    ws_steel['A3'].alignment = center_align

    ws_steel.merge_cells(f'A4:{last_steel_col}4')
    ws_steel['A4'] = "STEEL DETAILS"
    ws_steel['A4'].font = Font(bold=True, size=12)
    ws_steel['A4'].alignment = center_align

    ws_steel.append([])

    steel_headers = (
        ["Sr.No", "Description of Item","Location Of The Bar", "Nos", "Length", "Width", "Depth",
        "Ftg. Depth", "Dia Of Bar", "Spacing", "Bar Nos", "Bar Length","Column Height", "Top Beam Depth"]
        + [f"{d} MM" for d in BAR_DIAMETERS]
        + [ "Remark"]
    )
    steel_total_cols = len(steel_headers)  
    last_steel_col = get_column_letter(steel_total_cols)

    ws_steel.append(steel_headers)

    for col_num in range(1, len(steel_headers) + 1):
        cell = ws_steel.cell(row=6, column=col_num)
        cell.font = bold_font
        cell.alignment = center_align if col_num != 2 else left_align
        cell.border = thin_border

    # --- Build the nested tree: {id: {"subject":..., "children": {...}, "rows": [...]}}
    tree = {}
    node_order = []  # preserve first-seen order at each level

    def get_or_create_node(container, order_list, node_id, subject):
        if node_id not in container:
            container[node_id] = {"subject": subject, "children": {}, "children_order": [], "rows": []}
            order_list.append(node_id)
        return container[node_id]

    for row in getattr(doc, "ra_steel_details", []):
        path = get_row_hierarchy(row)
        if not path:
            continue

        node = get_or_create_node(tree, node_order, path[0][0], path[0][1])

        for node_id, subject in path[1:]:
            node = get_or_create_node(node["children"], node["children_order"], node_id, subject)

        node["rows"].append(row)

    # --- Recursively write the tree to the sheet, computing weight roll-ups ---
    row_num = 7
    sr_no = 1
    grand_total_weight = 0.0

    # stage-level (top-level) weight roll-up, needed by the Measurement sheet
    steel_stage_weights.clear()
    steel_stage_totals.clear()
    steel_task_totals.clear()

    diameter_length_totals = {d: 0.0 for d in BAR_DIAMETERS}
    def write_leaf_row(r, row_label, indent, is_first_row):
            """Writes one data row. If is_first_row, puts the task subject in Description too."""
            nonlocal row_num

            cell_sr = ws_steel.cell(row=row_num, column=1, value="")
            cell_sr.border = thin_border

            desc_val = (indent + row_label) if is_first_row else ""
            cell_desc = ws_steel.cell(row=row_num, column=2, value=desc_val)
            if is_first_row:
                cell_desc.font = bold_font
            cell_desc.border = thin_border
            cell_desc.alignment = left_align

            loc_cell = ws_steel.cell(row=row_num, column=3, value=r.get("location_of_the_bar") or "")
            loc_cell.border = thin_border
            loc_cell.alignment = left_align

            plain_fields = [
                ("nos", 4), ("length", 5), ("width", 6), ("depth", 7),
                ("ftg_depth", 8), ("dia_of_bar", 9), ("spacing", 10),
                ("bar_nos", 11), ("bar_length", 12),
                ("column_height", 13), ("top_beam_depth", 14)
            ]
            for fieldname, col in plain_fields:
                val = flt(r.get(fieldname))
                cell = ws_steel.cell(row=row_num, column=col, value=val if val else "")
                cell.border = thin_border
                cell.alignment = center_align
                if val:
                    cell.number_format = '0.00'

            reinforcement = get_row_reinforcement(r)
            for idx, d in enumerate(BAR_DIAMETERS):
                val = reinforcement.get(d, "")
                cell = ws_steel.cell(row=row_num, column=15 + idx, value=val)
                cell.border = thin_border
                cell.alignment = center_align
                if val:
                    cell.number_format = '0.00'
                    diameter_length_totals[d] += val   # accumulate for bottom summary

            remark_cell = ws_steel.cell(row=row_num, column=23, value=r.get("remark") or "")
            remark_cell.border = thin_border
            remark_cell.alignment = left_align

            row_num += 1
            return get_row_weight(r)

    def write_node(node, depth, is_top_level, stage_key=None, task_key=None):
        nonlocal row_num, sr_no, grand_total_weight

        indent = "  " * depth
        node_weight = 0.0

        is_leaf = not node["children_order"]
        if is_leaf and node["rows"]:
            sr_row = row_num

            node_weight += write_leaf_row(node["rows"][0], node["subject"], indent, is_first_row=True)

            cell_sr = ws_steel.cell(row=sr_row, column=1, value=sr_no)
            cell_sr.border = thin_border
            cell_sr.alignment = left_align
            sr_no += 1

            for r in node["rows"][1:]:
                node_weight += write_leaf_row(r, node["subject"], indent, is_first_row=False)

        else:
            cell_sr = ws_steel.cell(row=row_num, column=1, value=sr_no)
            cell_sr.border = thin_border
            cell_sr.alignment = left_align
            sr_no += 1

            cell_desc = ws_steel.cell(row=row_num, column=2, value=indent + node["subject"])
            cell_desc.font = bold_font
            cell_desc.border = thin_border
            cell_desc.alignment = left_align

            for c in range(3, steel_total_cols + 1):
                cell = ws_steel.cell(row=row_num, column=c, value="")
                cell.border = thin_border

            row_num += 1

            for r in node["rows"]:
                node_weight += write_leaf_row(r, "", indent, is_first_row=False)

        for child_id in node["children_order"]:
            child = node["children"][child_id]
            child_weight = write_node(
                child, depth + 1,
                is_top_level=False, stage_key=stage_key, task_key=task_key
            )
            node_weight += child_weight


        if node_weight and depth == 1 and task_key is not None:
            steel_task_totals[task_key] = steel_task_totals.get(task_key, 0) + node_weight

        return node_weight

    for stage_id in node_order:
        stage_node = tree[stage_id]
        stage_weight = write_node(stage_node, 0, True, stage_key=stage_id, task_key=stage_id)

        steel_stage_weights[stage_node["subject"]] = steel_stage_weights.get(stage_node["subject"], 0) + stage_weight
        steel_stage_totals[stage_node["subject"]] = steel_stage_weights[stage_node["subject"]] / 1000
        grand_total_weight += stage_weight

# --- Grand Total (Kg) ---
    # cell_desc = ws_steel.cell(row=row_num, column=2, value="GRAND TOTAL")
    # cell_desc.font = bold_font
    # cell_desc.border = thin_border
    # cell_desc.alignment = left_align

    # for c in [1] + list(range(3, steel_total_cols + 1)):
    #     cell = ws_steel.cell(row=row_num, column=c, value="")
    #     cell.border = thin_border

    # gt_cell = ws_steel.cell(row=row_num, column=3, value=f"{grand_total_weight:.2f} Kg")
    # gt_cell.font = bold_font
    # gt_cell.border = thin_border
    # gt_cell.alignment = left_align
    row_num += 1  # leave a blank spacer row before the summary block

    # --- Bottom Summary Block (Length / Weight Per Meter / Weight in Kg / Weight in MT / Total) ---
    diameter_col_start = 15  # matches reinforcement columns in the main table

    def write_summary_row(label, value_map, bold=False, number_format='0.00'):
        nonlocal row_num
        cell_label = ws_steel.cell(row=row_num, column=2, value=label)
        if bold:
            cell_label.font = bold_font
        cell_label.border = thin_border
        cell_label.alignment = left_align

        for c in [1] + list(range(3, diameter_col_start)):
            cell = ws_steel.cell(row=row_num, column=c, value="")
            cell.border = thin_border

        for idx, d in enumerate(BAR_DIAMETERS):
            val = value_map.get(d, 0)
            cell = ws_steel.cell(row=row_num, column=diameter_col_start + idx, value=round(val, 3) if val else 0)
            if bold:
                cell.font = bold_font
            cell.border = thin_border
            cell.alignment = center_align
            cell.number_format = number_format

        cell_remark = ws_steel.cell(row=row_num, column=23, value="")
        cell_remark.border = thin_border

        row_num += 1

    weight_kg_by_dia = {
        d: diameter_length_totals[d] * BAR_WEIGHT_PER_METER[d] for d in BAR_DIAMETERS
    }
    weight_mt_by_dia = {
        d: weight_kg_by_dia[d] / 1000 for d in BAR_DIAMETERS
    }

    write_summary_row("Length In Meter", diameter_length_totals)
    write_summary_row("Weight Per Meter", BAR_WEIGHT_PER_METER,number_format='0.000')
    write_summary_row("Weight In Kg", weight_kg_by_dia)
    write_summary_row("Weight In MT", weight_mt_by_dia, bold=True, number_format='0.000')

    # Single grand total line (sum across all diameters)
    total_mt = sum(weight_mt_by_dia.values())
    cell_label = ws_steel.cell(row=row_num, column=2, value="Total Weight In MT")
    cell_label.font = bold_font
    cell_label.border = thin_border
    cell_label.alignment = left_align

    for c in [1] + list(range(3, diameter_col_start)):
        cell = ws_steel.cell(row=row_num, column=c, value="")
        cell.border = thin_border

    total_cell = ws_steel.cell(row=row_num, column=diameter_col_start, value=round(total_mt, 3))
    total_cell.font = bold_font
    total_cell.border = thin_border
    total_cell.alignment = center_align
    total_cell.number_format = '0.000'

    for c in range(diameter_col_start + 1, 23):
        cell = ws_steel.cell(row=row_num, column=c, value="")
        cell.border = thin_border

    cell_remark = ws_steel.cell(row=row_num, column=23, value="")
    cell_remark.border = thin_border
    row_num += 1

    for col in ws_steel.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws_steel.column_dimensions[column].width = (max_length + 2)
    ws_steel.column_dimensions['A'].width = 8
    # # --- Steel Sheet ---
    # ws_steel = wb.create_sheet("Steel") 
    # ws_steel.merge_cells('A1:L1')
    # ws_steel['A1'] = company_name
    # ws_steel['A1'].font = Font(bold=True, size=14)
    # ws_steel['A1'].alignment = center_align

    # ws_steel.merge_cells('A2:L2')
    # ws_steel['A2'] = project_doc.project_name or project_doc.name
    # ws_steel['A2'].font = Font(bold=True, size=12)
    # ws_steel['A2'].alignment = center_align

    # ws_steel.merge_cells('A3:L3')
    # ws_steel['A3'] = f"RA Bill Number: {doc.name}"
    # ws_steel['A3'].font = Font(bold=True, size=12)
    # ws_steel['A3'].alignment = center_align

    # ws_steel.merge_cells('A4:L4')
    # ws_steel['A4'] = "STEEL DETAILS"
    # ws_steel['A4'].font = Font(bold=True, size=12)
    # ws_steel['A4'].alignment = center_align

    # ws_steel.append([])
    # steel_headers = ["Sr. No", "Content", "Description", "No of FDN", "No of Bar", "Dia meter of Bar", "Cutting Length", "Total Length", "Quantity", "Weight of Bar", "Total Weight", "Unit"]
    # ws_steel.append(steel_headers)
    
    # for col_num in range(1, 13):
    #     cell = ws_steel.cell(row=6, column=col_num)
    #     cell.font = bold_font
    #     cell.alignment = center_align if col_num not in (2, 3) else left_align
    #     cell.border = thin_border

    # grouped_steel_data = {}
    # for row in getattr(doc, "ra_steel_details", []):
    #     task = row.task
    #     subtask = row.subtask
        
    #     stage_subject = "No Stage"
    #     task_subject = "No Task"
    #     if task:
    #         task_info = frappe.db.get_value("Task", task, ["subject", "parent_task"], as_dict=True)
    #         if task_info:
    #             task_subject = task_info.get("subject") or "No Task"
    #             stage_id = task_info.get("parent_task")
    #             if stage_id:
    #                 stage_subj = frappe.db.get_value("Task", stage_id, "subject")
    #                 if stage_subj:
    #                     stage_subject = stage_subj
        
    #     subtask_subject = "No Subtask"
    #     if subtask:
    #         subtask_subj = frappe.db.get_value("Task", subtask, "subject")
    #         if subtask_subj:
    #             subtask_subject = subtask_subj
        
    #     if stage_subject not in grouped_steel_data:
    #         grouped_steel_data[stage_subject] = {}
    #     if task_subject not in grouped_steel_data[stage_subject]:
    #         grouped_steel_data[stage_subject][task_subject] = {}
    #     if subtask_subject not in grouped_steel_data[stage_subject][task_subject]:
    #         grouped_steel_data[stage_subject][task_subject][subtask_subject] = []
            
    #     grouped_steel_data[stage_subject][task_subject][subtask_subject].append(row)
        
    # row_num = 7
    # stage_idx = 1
    # grand_total_weight = 0.0
    # for stage, tasks in grouped_steel_data.items():
    #     stage_total_weight = 0.0
    #     cell_sr = ws_steel.cell(row=row_num, column=1, value=str(stage_idx))
    #     cell_sr.border = thin_border; cell_sr.alignment = left_align
        
    #     cell_desc = ws_steel.cell(row=row_num, column=2, value=stage)
    #     cell_desc.font = bold_font; cell_desc.border = thin_border; cell_desc.alignment = left_align
        
    #     for c in range(3, 13):
    #         cell = ws_steel.cell(row=row_num, column=c, value="")
    #         cell.border = thin_border; cell.alignment = center_align
    #     row_num += 1
        
    #     task_idx = 1
    #     for task, subtasks in tasks.items():
    #         cell_sr = ws_steel.cell(row=row_num, column=1, value=f"{stage_idx}.{task_idx}")
    #         cell_sr.border = thin_border; cell_sr.alignment = left_align
            
    #         cell_desc = ws_steel.cell(row=row_num, column=2, value="  " + task)
    #         cell_desc.font = bold_font; cell_desc.border = thin_border; cell_desc.alignment = left_align
            
    #         for c in range(3, 13):
    #             cell = ws_steel.cell(row=row_num, column=c, value="")
    #             cell.border = thin_border; cell.alignment = center_align
    #         row_num += 1
            
    #         subtask_idx = 1
    #         task_total_weight = 0.0
    #         for subtask, rows in subtasks.items():
    #             subtask_id = rows[0].subtask if len(rows) > 0 else None
    #             task_desc = get_task_desc(subtask_id)
                
    #             for r in rows:
    #                 cell_sr = ws_steel.cell(row=row_num, column=1, value=f"{stage_idx}.{task_idx}.{subtask_idx}")
    #                 cell_sr.border = thin_border; cell_sr.alignment = left_align
                    
    #                 cell_desc = ws_steel.cell(row=row_num, column=2, value="    " + subtask)
    #                 cell_desc.border = thin_border; cell_desc.alignment = left_align
                    
    #                 # If r.description exists, use it. Else use task_desc.
    #                 # final_desc = r.description if r.description else task_desc
    #                 # cell_subdesc = ws_steel.cell(row=row_num, column=3, value=final_desc)
    #                 cell_subdesc.border = thin_border; cell_subdesc.alignment = left_align
                    
    #                 cell = ws_steel.cell(row=row_num, column=4, value=r.no_of_fdn or "")
    #                 cell.border = thin_border; cell.alignment = center_align
                    
    #                 cell = ws_steel.cell(row=row_num, column=5, value=r.no_of_bar or "")
    #                 cell.border = thin_border; cell.alignment = center_align
                    
    #                 cell = ws_steel.cell(row=row_num, column=6, value=r.diamter_of_bar or "")
    #                 cell.border = thin_border; cell.alignment = center_align
                    
    #                 cell = ws_steel.cell(row=row_num, column=7, value=r.cutting_length or "")
    #                 cell.border = thin_border; cell.alignment = center_align
                    
    #                 t_len = flt(r.total_length)
    #                 cell = ws_steel.cell(row=row_num, column=8, value=t_len)
    #                 cell.border = thin_border; cell.alignment = center_align
    #                 if t_len: cell.number_format = '0.00'
                    
    #                 qty = flt(r.qty)
    #                 cell = ws_steel.cell(row=row_num, column=9, value=qty)
    #                 cell.border = thin_border; cell.alignment = center_align
    #                 if qty: cell.number_format = '0.00'
                    
    #                 w_bar = flt(r.weight_of_bar)
    #                 cell = ws_steel.cell(row=row_num, column=10, value=w_bar)
    #                 cell.border = thin_border; cell.alignment = center_align
    #                 if w_bar: cell.number_format = '0.00'
                    
    #                 t_weight = flt(r.total_weight)
    #                 task_total_weight += t_weight
    #                 grand_total_weight += t_weight
    #                 cell = ws_steel.cell(row=row_num, column=11, value=t_weight)
    #                 cell.border = thin_border; cell.alignment = center_align
    #                 if t_weight: cell.number_format = '0.00'
                    
    #                 cell = ws_steel.cell(row=row_num, column=12, value=r.unit or "")
    #                 cell.border = thin_border; cell.alignment = center_align
                    
    #                 row_num += 1
    #             subtask_idx += 1
    #         cell_sr = ws_steel.cell(row=row_num, column=1, value="")
    #         cell_sr.border = thin_border
    #         cell_sr.alignment = left_align

    #         cell_desc = ws_steel.cell(
    #             row=row_num,
    #             column=2,
    #             value="    Total"
    #         )
    #         cell_desc.font = bold_font
    #         cell_desc.border = thin_border
    #         cell_desc.alignment = left_align

    #         for c in range(3, 11):
    #             cell = ws_steel.cell(row=row_num, column=c, value="")
    #             cell.border = thin_border
    #             cell.alignment = center_align

    #         total_weight_cell = ws_steel.cell(
    #             row=row_num,
    #             column=11,
    #             value=task_total_weight
    #         )
    #         total_weight_cell.font = bold_font
    #         total_weight_cell.border = thin_border
    #         total_weight_cell.alignment = center_align
    #         total_weight_cell.number_format = '0.00'
    #         steel_task_totals[f"{stage}||{task}"] = task_total_weight
    #         stage_total_weight += task_total_weight
    #         cell = ws_steel.cell(row=row_num, column=12, value="")
    #         cell.border = thin_border
    #         cell.alignment = center_align

    #         row_num += 1
    #         # Metric Tonne Conversion Row
    #         cell_sr = ws_steel.cell(row=row_num, column=1, value="")
    #         cell_sr.border = thin_border

    #         cell_desc = ws_steel.cell(
    #             row=row_num,
    #             column=2,
    #             value="    Total (Metric Tonne)"
    #         )
    #         cell_desc.font = bold_font
    #         cell_desc.border = thin_border
    #         cell_desc.alignment = left_align

    #         for c in range(3, 11):
    #             cell = ws_steel.cell(row=row_num, column=c, value="")
    #             cell.border = thin_border
    #             cell.alignment = center_align

    #         mt_cell = ws_steel.cell(
    #             row=row_num,
    #             column=11,
    #             value=(task_total_weight / 1000)
    #         )
    #         mt_cell.font = bold_font
    #         mt_cell.border = thin_border
    #         mt_cell.alignment = center_align
    #         mt_cell.number_format = '0.000'

    #         unit_cell = ws_steel.cell(
    #             row=row_num,
    #             column=12,
    #             value="Metric Tonne"
    #         )
    #         unit_cell.border = thin_border
    #         unit_cell.alignment = center_align

    #         row_num += 1

    #         task_idx += 1
    #         steel_stage_weights[stage] = stage_total_weight
    #         steel_stage_totals[stage] = stage_total_weight / 1000
    #     stage_idx += 1
    #     stage_idx += 1

    # cell_sr = ws_steel.cell(row=row_num, column=1, value="")
    # cell_sr.border = thin_border

    # cell_desc = ws_steel.cell(
    #     row=row_num,
    #     column=2,
    #     value="GRAND TOTAL"
    # )
    # cell_desc.font = bold_font
    # cell_desc.border = thin_border
    # cell_desc.alignment = left_align

    # for c in range(3, 11):
    #     cell = ws_steel.cell(row=row_num, column=c, value="")
    #     cell.border = thin_border
    #     cell.alignment = center_align

    # gt_cell = ws_steel.cell(
    #     row=row_num,
    #     column=11,
    #     value=grand_total_weight
    # )
    # gt_cell.font = bold_font
    # gt_cell.border = thin_border
    # gt_cell.alignment = center_align
    # gt_cell.number_format = '0.00'

    # cell = ws_steel.cell(row=row_num, column=12, value="Kg")
    # cell.border = thin_border
    # cell.alignment = center_align

    # row_num += 1

    # cell_sr = ws_steel.cell(row=row_num, column=1, value="")
    # cell_sr.border = thin_border

    # cell_desc = ws_steel.cell(
    #     row=row_num,
    #     column=2,
    #     value="GRAND TOTAL"
    # )
    # cell_desc.font = bold_font
    # cell_desc.border = thin_border
    # cell_desc.alignment = left_align

    # for c in range(3, 11):
    #     cell = ws_steel.cell(row=row_num, column=c, value="")
    #     cell.border = thin_border
    #     cell.alignment = center_align

    # mt_cell = ws_steel.cell(
    #     row=row_num,
    #     column=11,
    #     value=(grand_total_weight / 1000)
    # )
    # mt_cell.font = bold_font
    # mt_cell.border = thin_border
    # mt_cell.alignment = center_align
    # mt_cell.number_format = '0.000'

    # cell = ws_steel.cell(row=row_num, column=12, value="Metric Tonne")
    # cell.border = thin_border
    # cell.alignment = center_align
    # ws_steel.column_dimensions[get_column_letter(12)].width = 15
    # for col in ws_steel.columns:
    #     max_length = 0
    #     column = get_column_letter(col[0].column)
    #     for cell in col:
    #         try:
    #             if len(str(cell.value)) > max_length:
    #                 max_length = len(str(cell.value))
    #         except: pass
    #     ws_steel.column_dimensions[column].width = (max_length + 2)

    # --- Measurement Sheet ---
    # ws_meas.merge_cells('A1:J1')
    # ws_meas['A1'] = company_name
    # ws_meas['A1'].font = Font(bold=True, size=14)
    # ws_meas['A1'].alignment = center_align

    # ws_meas.merge_cells('A2:J2')
    # ws_meas['A2'] = project_doc.project_name or project_doc.name
    # ws_meas['A2'].font = Font(bold=True, size=12)
    # ws_meas['A2'].alignment = center_align

    # ws_meas.merge_cells('A3:J3')
    # ws_meas['A3'] = f"RA Bill Number: {doc.name}"
    # ws_meas['A3'].font = Font(bold=True, size=12)
    # ws_meas['A3'].alignment = center_align

    # ws_meas.merge_cells('A4:J4')
    # ws_meas['A4'] = "MEASUREMENT SHEET"
    # ws_meas['A4'].font = Font(bold=True, size=12)
    # ws_meas['A4'].alignment = center_align

    # ws_meas.append([])
    # meas_headers = ["Sr. No", "Content", "Description", "No.1", "No.2", "Length", "Width", "Depth", "Unit", "Qty"]
    # ws_meas.append(meas_headers)

    # for col_num in range(1, 11):
    #     cell = ws_meas.cell(row=6, column=col_num)
    #     cell.font = bold_font
    #     cell.alignment = center_align if col_num not in (2, 3) else left_align
    #     cell.border = thin_border

    # steel_subtask_weights = {}
    # for steel_row in getattr(doc, "ra_steel_details", []):
    #     path = get_row_hierarchy(steel_row)
    #     deepest_id = path[-1][0] if path else None

    #     key = (steel_row.get("task"), deepest_id)

    #     steel_subtask_weights[key] = (
    #         steel_subtask_weights.get(key, 0)
    #         + get_row_weight(steel_row)
    #     )

    # grouped_meas_data = {}
    # for row in getattr(doc, "ra_billing_details", []):
    #     stage = row.stage_subject or "No Stage"
    #     task = row.task_subject or "No Task"
    #     if stage not in grouped_meas_data:
    #         grouped_meas_data[stage] = {}
    #     if task not in grouped_meas_data[stage]:
    #         grouped_meas_data[stage][task] = []
    #     grouped_meas_data[stage][task].append(row)
        
    # row_num = 7
    # stage_idx = 1
    # for stage, tasks in grouped_meas_data.items():
    #     current_stage = stage
    #     cell_sr = ws_meas.cell(row=row_num, column=1, value=str(stage_idx))
    #     cell_sr.border = thin_border; cell_sr.alignment = left_align
        
    #     cell_desc = ws_meas.cell(row=row_num, column=2, value=stage)
    #     cell_desc.font = bold_font
    #     cell_desc.border = thin_border
    #     cell_desc.alignment = left_align
        
    #     for c in range(3, 11):
    #         cell = ws_meas.cell(row=row_num, column=c, value="")
    #         cell.border = thin_border; cell.alignment = center_align
            
    #     row_num += 1
        
    #     task_idx = 1
    #     show_mt_total = False
    #     for task, subtasks in tasks.items():
    #         current_task = task
    #         cell_sr = ws_meas.cell(row=row_num, column=1, value=f"{stage_idx}.{task_idx}")
    #         cell_sr.border = thin_border; cell_sr.alignment = left_align
            
    #         cell_desc = ws_meas.cell(row=row_num, column=2, value="  " + task)
    #         cell_desc.font = bold_font
    #         cell_desc.border = thin_border
    #         cell_desc.alignment = left_align
            
    #         for c in range(3, 11):
    #             cell = ws_meas.cell(row=row_num, column=c, value="")
    #             cell.border = thin_border; cell.alignment = center_align
            
    #         row_num += 1
            
    #         subtask_idx = 1
    #         task_total_qty = 0.0
            
    #         for row in subtasks:
    #             cell_sr = ws_meas.cell(row=row_num, column=1, value=f"{stage_idx}.{task_idx}.{subtask_idx}")
    #             cell_sr.border = thin_border; cell_sr.alignment = left_align
                
    #             cell_desc = ws_meas.cell(row=row_num, column=2, value="    " + (row.subtask_subject or ""))
    #             cell_desc.border = thin_border; cell_desc.alignment = left_align
                
    #             desc_val = row.description if hasattr(row, 'description') and row.description else get_task_desc(row.subtask)
    #             cell_subdesc = ws_meas.cell(row=row_num, column=3, value=desc_val)
    #             cell_subdesc.border = thin_border; cell_subdesc.alignment = left_align
                
    #             cell = ws_meas.cell(row=row_num, column=4, value=row.no1 or "")
    #             cell.border = thin_border; cell.alignment = center_align
                
    #             cell = ws_meas.cell(row=row_num, column=5, value=row.no2 or "")
    #             cell.border = thin_border; cell.alignment = center_align
                
    #             cell = ws_meas.cell(row=row_num, column=6, value=row.length or "")
    #             cell.border = thin_border; cell.alignment = center_align
                
    #             cell = ws_meas.cell(row=row_num, column=7, value=row.width or "")
    #             cell.border = thin_border; cell.alignment = center_align
                
    #             cell = ws_meas.cell(row=row_num, column=8, value=row.height or "")
    #             cell.border = thin_border; cell.alignment = center_align
                
    #             cell = ws_meas.cell(row=row_num, column=9, value=row.uom or "")
    #             if (row.uom or "").lower() in ["kg", "kilogram"]:
    #                 show_mt_total = True
    #             cell.border = thin_border; cell.alignment = center_align
    #             #unit cell
    #             unit_cell = ws_meas.cell(
    #                 row=row_num,
    #                 column=9,
    #                 value=row.uom or ""
    #             )
    #             unit_cell.border = thin_border
    #             unit_cell.alignment = center_align
    #             #qty cell
    #             qty = flt(row.quantity)
    #             task_total_qty += qty
    #             steel_weight = steel_subtask_weights.get(
    #                 (row.task, row.subtask),
    #                 0
    #             )
    #             # qty_cell = ws_meas.cell(row=row_num, column=10, value=steel_weight)

    #             if (row.uom or "").lower() in ["kg", "kilogram"]:
    #                 display_qty = steel_subtask_weights.get(
    #                     (row.task, row.subtask),
    #                     qty
    #                 )
    #             else:
    #                 display_qty = qty
    #             qty_cell = ws_meas.cell(row=row_num, column=10, value=display_qty)
    #             qty_cell.border = thin_border; qty_cell.alignment = center_align
    #             qty_cell.number_format = '0.00'
                
    #             row_num += 1
    #             subtask_idx += 1
                
    #         task_idx += 1
    #     if show_mt_total:

    #         stage_kg_total = steel_stage_weights.get(stage, 0)

    #         cell_sr = ws_meas.cell(row=row_num, column=1, value="")
    #         cell_sr.border = thin_border

    #         cell_desc = ws_meas.cell(
    #             row=row_num,
    #             column=2,
    #             value="Stage Total (Kg)"
    #         )
    #         cell_desc.font = bold_font
    #         cell_desc.border = thin_border
    #         cell_desc.alignment = left_align

    #         for c in range(3, 10):
    #             cell = ws_meas.cell(row=row_num, column=c, value="")
    #             cell.border = thin_border
    #             cell.alignment = center_align

    #         kg_cell = ws_meas.cell(
    #             row=row_num,
    #             column=10,
    #             value=stage_kg_total
    #         )

    #         kg_cell.font = bold_font
    #         kg_cell.border = thin_border
    #         kg_cell.alignment = center_align
    #         kg_cell.number_format = "0.00"

    #         row_num += 1
    #         stage_mt_total = steel_stage_totals.get(stage, 0)

    #         cell_sr = ws_meas.cell(row=row_num, column=1, value="")
    #         cell_sr.border = thin_border

    #         cell_desc = ws_meas.cell(
    #             row=row_num,
    #             column=2,
    #             value="Stage Total (Metric Tonne)"
    #         )
    #         cell_desc.font = bold_font
    #         cell_desc.border = thin_border
    #         cell_desc.alignment = left_align

    #         for c in range(3, 10):
    #             cell = ws_meas.cell(row=row_num, column=c, value="")
    #             cell.border = thin_border
    #             cell.alignment = center_align

    #         mt_cell = ws_meas.cell(
    #             row=row_num,
    #             column=10,
    #             value=stage_mt_total
    #         )

    #         mt_cell.font = bold_font
    #         mt_cell.border = thin_border
    #         mt_cell.alignment = center_align
    #         mt_cell.number_format = "0.000"
    #         row_num += 1

    #     stage_idx += 1

    # for col in ws_meas.columns:
    #     max_length = 0
    #     column = get_column_letter(col[0].column)
    #     for cell in col:
    #         try:
    #             if len(str(cell.value)) > max_length:
    #                 max_length = len(str(cell.value))
    #         except: pass
    #     ws_meas.column_dimensions[column].width = (max_length + 2)

    # --- Measurement Sheet ---
    ws_meas.merge_cells('A1:J1')
    ws_meas['A1'] = company_name
    ws_meas['A1'].font = Font(bold=True, size=14)
    ws_meas['A1'].alignment = center_align

    ws_meas.merge_cells('A2:J2')
    ws_meas['A2'] = project_doc.project_name or project_doc.name
    ws_meas['A2'].font = Font(bold=True, size=12)
    ws_meas['A2'].alignment = center_align

    ws_meas.merge_cells('A3:J3')
    ws_meas['A3'] = f"RA Bill Number: {doc.name}"
    ws_meas['A3'].font = Font(bold=True, size=12)
    ws_meas['A3'].alignment = center_align

    ws_meas.merge_cells('A4:J4')
    ws_meas['A4'] = "MEASUREMENT SHEET"
    ws_meas['A4'].font = Font(bold=True, size=12)
    ws_meas['A4'].alignment = center_align

    ws_meas.append([])
    meas_headers = ["Sr. No", "Content", "Description", "No.1", "No.2", "Length", "Width", "Depth", "Unit", "Qty"]
    ws_meas.append(meas_headers)

    for col_num in range(1, 11):
        cell = ws_meas.cell(row=6, column=col_num)
        cell.font = bold_font
        cell.alignment = center_align if col_num not in (2, 3) else left_align
        cell.border = thin_border

    # --- deepest-task-id helper for ra_billing_details rows (mirrors update_abstract_details logic) ---
    BILLING_LEVEL_FIELDS = [f"task_level{i}" for i in range(1, 11)]
    BILLING_LEVEL_SUBJECT_FIELDS = [f"level{i}_subject" for i in range(1, 11)]

    def get_deepest_billing_task_id(row):
        for lvl_field in reversed(BILLING_LEVEL_FIELDS):
            if row.get(lvl_field):
                return row.get(lvl_field)
        return row.get("task")

    def get_billing_row_hierarchy(row):
        """
        Ordered list of (id, subject): stage, task, then whichever
        task_level1..10 are actually populated on this row.
        """
        path = []
        if row.get("stage"):
            path.append((row.get("stage"), row.get("stage_subject") or row.get("stage")))
        if row.get("task"):
            path.append((row.get("task"), row.get("task_subject") or row.get("task")))
        for lvl_field, subj_field in zip(BILLING_LEVEL_FIELDS, BILLING_LEVEL_SUBJECT_FIELDS):
            if row.get(lvl_field):
                path.append((row.get(lvl_field), row.get(subj_field) or row.get(lvl_field)))
        return path

    # steel weight lookup keyed by (task, deepest_id) — matches how it's built in the Steel sheet section
    steel_subtask_weights = {}
    for steel_row in getattr(doc, "ra_steel_details", []):
        path = get_row_hierarchy(steel_row)
        deepest_id = path[-1][0] if path else None
        key = (steel_row.get("task"), deepest_id)
        steel_subtask_weights[key] = (
            steel_subtask_weights.get(key, 0) + get_row_weight(steel_row)
        )

    # Which stages contain at least one Kg-uom row (drives the Kg/MT stage summary rows)
    stage_has_kg = {}
    for row in getattr(doc, "ra_billing_details", []):
        stage_subj = row.stage_subject or "No Stage"
        if (row.uom or "").lower() in ["kg", "kilogram"]:
            stage_has_kg[stage_subj] = True

    # --- Build the nested tree (same shape/helper as the Steel sheet) ---
    tree_meas = {}
    node_order_meas = []
    measurement_stage_kg_totals = {}

    for row in getattr(doc, "ra_billing_details", []):
        path = get_billing_row_hierarchy(row)
        if not path:
            continue

        node = get_or_create_node(tree_meas, node_order_meas, path[0][0], path[0][1])
        for node_id, subject in path[1:]:
            node = get_or_create_node(node["children"], node["children_order"], node_id, subject)

        node["rows"].append(row)

    row_num = 7
    sr_no=1

    def write_meas_leaf_row(r, row_label, indent, is_first_row):
        nonlocal row_num

        cell_sr = ws_meas.cell(row=row_num, column=1, value="")
        cell_sr.border = thin_border

        desc_val = (indent + row_label) if is_first_row else ""
        cell_desc = ws_meas.cell(row=row_num, column=2, value=desc_val)
        if is_first_row:
            cell_desc.font = bold_font
        cell_desc.border = thin_border
        cell_desc.alignment = left_align

        desc_text = r.description if getattr(r, "description", None) else get_task_desc(get_deepest_billing_task_id(r))
        cell_subdesc = ws_meas.cell(row=row_num, column=3, value=desc_text if is_first_row else "")
        cell_subdesc.border = thin_border
        cell_subdesc.alignment = left_align

        cell = ws_meas.cell(row=row_num, column=4, value=r.no1 or "")
        cell.border = thin_border; cell.alignment = center_align

        cell = ws_meas.cell(row=row_num, column=5, value=r.no2 or "")
        cell.border = thin_border; cell.alignment = center_align

        cell = ws_meas.cell(row=row_num, column=6, value=r.length or "")
        cell.border = thin_border; cell.alignment = center_align

        cell = ws_meas.cell(row=row_num, column=7, value=r.width or "")
        cell.border = thin_border; cell.alignment = center_align

        cell = ws_meas.cell(row=row_num, column=8, value=r.height or "")
        cell.border = thin_border; cell.alignment = center_align

        uom_val = r.uom or ""
        cell = ws_meas.cell(row=row_num, column=9, value=uom_val)
        cell.border = thin_border; cell.alignment = center_align

        qty = flt(r.quantity)
        if uom_val.lower() in ["kg", "kilogram"]:
            deepest_id = get_deepest_billing_task_id(r)
            display_qty = steel_subtask_weights.get((r.task, deepest_id), qty)

            # ACCUMULATE per-stage sum from what's actually shown in this sheet
            stage_key = r.stage_subject or "No Stage"
            measurement_stage_kg_totals[stage_key] = (
                measurement_stage_kg_totals.get(stage_key, 0) + display_qty
            )
        else:
            display_qty = qty

        qty_cell = ws_meas.cell(row=row_num, column=10, value=display_qty)
        qty_cell.border = thin_border; qty_cell.alignment = center_align
        qty_cell.number_format = '0.00'

        row_num += 1

    def write_meas_node(node, depth):
        nonlocal row_num, sr_no

        indent = "  " * depth
        is_leaf = not node["children_order"]

        if is_leaf and node["rows"]:
            sr_row = row_num

            write_meas_leaf_row(node["rows"][0], node["subject"], indent, is_first_row=True)

            cell_sr = ws_meas.cell(row=sr_row, column=1, value=sr_no)
            cell_sr.border = thin_border; cell_sr.alignment = left_align
            sr_no += 1

            for r in node["rows"][1:]:
                write_meas_leaf_row(r, "", indent, is_first_row=False)

        else:
            cell_sr = ws_meas.cell(row=row_num, column=1, value=sr_no)
            cell_sr.border = thin_border; cell_sr.alignment = left_align
            sr_no += 1

            cell_desc = ws_meas.cell(row=row_num, column=2, value=indent + node["subject"])
            cell_desc.font = bold_font
            cell_desc.border = thin_border
            cell_desc.alignment = left_align

            for c in range(3, 11):
                cell = ws_meas.cell(row=row_num, column=c, value="")
                cell.border = thin_border

            row_num += 1

            for r in node["rows"]:
                write_meas_leaf_row(r, "", indent, is_first_row=False)

        child_idx = 1
        for child_id in node["children_order"]:
            child = node["children"][child_id]
            write_meas_node(child, depth + 1)

    for stage_id in node_order_meas:
        stage_node = tree_meas[stage_id]
        write_meas_node(stage_node, 0)

        if stage_has_kg.get(stage_node["subject"]):
            stage_kg_total = measurement_stage_kg_totals.get(stage_node["subject"], 0)  
            stage_mt_total = stage_kg_total / 1000                                        

            cell_sr = ws_meas.cell(row=row_num, column=1, value="")
            cell_sr.border = thin_border

            cell_desc = ws_meas.cell(row=row_num, column=2, value="Stage Total (Kg)")
            cell_desc.font = bold_font; cell_desc.border = thin_border; cell_desc.alignment = left_align

            for c in range(3, 10):
                cell = ws_meas.cell(row=row_num, column=c, value="")
                cell.border = thin_border; cell.alignment = center_align

            kg_cell = ws_meas.cell(row=row_num, column=10, value=stage_kg_total)
            kg_cell.font = bold_font; kg_cell.border = thin_border
            kg_cell.alignment = center_align; kg_cell.number_format = "0.00"
            row_num += 1

            cell_sr = ws_meas.cell(row=row_num, column=1, value="")
            cell_sr.border = thin_border

            cell_desc = ws_meas.cell(row=row_num, column=2, value="Stage Total (Metric Tonne)")
            cell_desc.font = bold_font; cell_desc.border = thin_border; cell_desc.alignment = left_align

            for c in range(3, 10):
                cell = ws_meas.cell(row=row_num, column=c, value="")
                cell.border = thin_border; cell.alignment = center_align

            mt_cell = ws_meas.cell(row=row_num, column=10, value=stage_mt_total)
            mt_cell.font = bold_font; mt_cell.border = thin_border
            mt_cell.alignment = center_align; mt_cell.number_format = "0.000"
            row_num += 1

    measurement_stage_mt_totals = {
        stage: kg / 1000 for stage, kg in measurement_stage_kg_totals.items()
    }
    for col in ws_meas.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        ws_meas.column_dimensions[column].width = (max_length + 2)
    ws_meas.column_dimensions['A'].width = 8 
  
# --- Abstract Sheet ---
    abs_headers = [
        "Sr. No", "Stage", "Description", "UOM", "Rate",
        "Previous Bill Qty", "This Bill Qty", "Total Bill Qty",
        "Previous Bill Amount", "This Bill Amount", "Total Bill Amount",
        "Remarks"
    ]
    abs_total_cols = len(abs_headers)
    abs_last_col = get_column_letter(abs_total_cols)

    ws_abstract.merge_cells(f'A1:{abs_last_col}1')
    ws_abstract['A1'] = company_name
    ws_abstract['A1'].font = Font(bold=True, size=14)
    ws_abstract['A1'].alignment = center_align

    ws_abstract.merge_cells(f'A2:{abs_last_col}2')
    ws_abstract['A2'] = project_doc.project_name or project_doc.name
    ws_abstract['A2'].font = Font(bold=True, size=12)
    ws_abstract['A2'].alignment = center_align

    ws_abstract.merge_cells(f'A3:{abs_last_col}3')
    ws_abstract['A3'] = f"RA Bill Number: {doc.name}"
    ws_abstract['A3'].font = Font(bold=True, size=12)
    ws_abstract['A3'].alignment = center_align

    ws_abstract.merge_cells(f'A4:{abs_last_col}4')
    ws_abstract['A4'] = "ABSTRACT SHEET"
    ws_abstract['A4'].font = Font(bold=True, size=12)
    ws_abstract['A4'].alignment = center_align

    ws_abstract.append([])            # blank row -> row 5
    ws_abstract.append(abs_headers)   # header row -> row 6

    for col_num in range(1, abs_total_cols + 1):
        cell = ws_abstract.cell(row=6, column=col_num)
        cell.font = bold_font
        cell.alignment = center_align if col_num not in (2, 3, 12) else left_align
        cell.border = thin_border

    row_num = 7
    sr_no = 1

    grand_this_amt = 0.0
    grand_total_amt = 0.0

    for row in getattr(doc, "ra_abstract_details", []):

        stage_steel_kg = measurement_stage_kg_totals.get(row.stage_subject, 0)
        stage_steel_mt = measurement_stage_mt_totals.get(row.stage_subject, 0)
      
        rate = flt(row.rate)

        if stage_steel_kg:
            display_uom = "Metric Tonne"
            display_prev_qty = flt(row.previous_bill_quantity)
            display_this_qty = stage_steel_mt
            display_this_amt = rate * display_this_qty
        else:
            display_uom = row.uom or ""
            display_prev_qty = flt(row.previous_bill_quantity)
            display_this_qty = flt(row.billed_quantity)
            display_this_amt = flt(row.amount)

        display_total_qty = display_prev_qty + display_this_qty

        display_prev_amt = flt(row.previous_bill_amount)
        display_total_amt = display_prev_amt + display_this_amt

        cell_sr = ws_abstract.cell(row=row_num, column=1, value=sr_no)
        cell_sr.border = thin_border
        cell_sr.alignment = left_align

        cell_stage = ws_abstract.cell(row=row_num, column=2, value=row.stage_subject or "")
        cell_stage.font = bold_font
        cell_stage.border = thin_border
        cell_stage.alignment = left_align

        cell_desc = ws_abstract.cell(row=row_num, column=3, value=row.description or "")
        cell_desc.border = thin_border
        cell_desc.alignment = left_align

        cell_uom = ws_abstract.cell(row=row_num, column=4, value=display_uom)
        cell_uom.border = thin_border
        cell_uom.alignment = center_align

        rate_cell = ws_abstract.cell(row=row_num, column=5, value=rate)
        rate_cell.border = thin_border
        rate_cell.alignment = center_align
        rate_cell.number_format = '0.00'

        prev_qty_cell = ws_abstract.cell(row=row_num, column=6, value=display_prev_qty)
        prev_qty_cell.border = thin_border
        prev_qty_cell.alignment = center_align
        prev_qty_cell.number_format = '0.000' if stage_steel_kg else '0.00'

        this_qty_cell = ws_abstract.cell(row=row_num, column=7, value=display_this_qty)
        this_qty_cell.border = thin_border
        this_qty_cell.alignment = center_align
        this_qty_cell.number_format = '0.000' if stage_steel_kg else '0.00'

        total_qty_cell = ws_abstract.cell(row=row_num, column=8, value=display_total_qty)
        total_qty_cell.border = thin_border
        total_qty_cell.alignment = center_align
        total_qty_cell.number_format = '0.000' if stage_steel_kg else '0.00'

        prev_amt_cell = ws_abstract.cell(row=row_num, column=9, value=display_prev_amt)
        prev_amt_cell.border = thin_border
        prev_amt_cell.alignment = center_align
        prev_amt_cell.number_format = '0.00'

        this_amt_cell = ws_abstract.cell(row=row_num, column=10, value=display_this_amt)
        this_amt_cell.border = thin_border
        this_amt_cell.alignment = center_align
        this_amt_cell.number_format = '0.00'

        total_amt_cell = ws_abstract.cell(row=row_num, column=11, value=display_total_amt)
        total_amt_cell.border = thin_border
        total_amt_cell.alignment = center_align
        total_amt_cell.number_format = '0.00'

        remarks_cell = ws_abstract.cell(row=row_num, column=12, value=getattr(row, "remarks", "") or "")
        remarks_cell.border = thin_border
        remarks_cell.alignment = left_align

        grand_this_amt += display_this_amt
        grand_total_amt += display_total_amt

        sr_no += 1
        row_num += 1

    # --- Grand Total row: only This Bill Amount and Total Bill Amount populated ---
    cell_sr = ws_abstract.cell(row=row_num, column=1, value="")
    cell_sr.border = thin_border

    cell_desc = ws_abstract.cell(row=row_num, column=2, value="GRAND TOTAL")
    cell_desc.font = bold_font
    cell_desc.border = thin_border
    cell_desc.alignment = left_align

    for c in (3, 4, 5, 6, 7, 8, 9):
        cell = ws_abstract.cell(row=row_num, column=c, value="")
        cell.border = thin_border

    gt_this_amt = ws_abstract.cell(row=row_num, column=10, value=grand_this_amt)
    gt_this_amt.font = bold_font
    gt_this_amt.border = thin_border
    gt_this_amt.alignment = center_align
    gt_this_amt.number_format = '0.00'

    gt_total_amt = ws_abstract.cell(row=row_num, column=11, value=grand_total_amt)
    gt_total_amt.font = bold_font
    gt_total_amt.border = thin_border
    gt_total_amt.alignment = center_align
    gt_total_amt.number_format = '0.00'

    cell_remark = ws_abstract.cell(row=row_num, column=12, value="")
    cell_remark.border = thin_border

    row_num += 1

    for col in ws_abstract.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws_abstract.column_dimensions[column].width = (max_length + 2)
    ws_abstract.column_dimensions['A'].width = 8

    # --- Level Details Sheet ---
    ws_level = wb.create_sheet("Level Details")

    level_headers = [
        "Sr. No", "Task", "Task Subject",
        "Design", "B.S.", "I.S.", "F.S.", "H.I.", "R.L.",
        "Average R.L.", "Remark"
    ]
    total_level_cols = len(level_headers)
    last_level_col = get_column_letter(total_level_cols)

    ws_level.merge_cells(f'A1:{last_level_col}1')
    ws_level['A1'] = company_name
    ws_level['A1'].font = Font(bold=True, size=14)
    ws_level['A1'].alignment = center_align

    ws_level.merge_cells(f'A2:{last_level_col}2')
    ws_level['A2'] = project_doc.project_name or project_doc.name
    ws_level['A2'].font = Font(bold=True, size=12)
    ws_level['A2'].alignment = center_align

    ws_level.merge_cells(f'A3:{last_level_col}3')
    ws_level['A3'] = f"RA Bill Number: {doc.name}"
    ws_level['A3'].font = Font(bold=True, size=12)
    ws_level['A3'].alignment = center_align

    ws_level.merge_cells(f'A4:{last_level_col}4')
    ws_level['A4'] = "LEVEL DETAILS"
    ws_level['A4'].font = Font(bold=True, size=12)
    ws_level['A4'].alignment = center_align

    ws_level.append([])
    ws_level.append(level_headers)

    for col_num in range(1, total_level_cols + 1):
        cell = ws_level.cell(row=6, column=col_num)
        cell.font = bold_font
        cell.alignment = center_align if col_num not in (2, 3, 11) else left_align
        cell.border = thin_border

    row_num = 7
    sr_no = 1

    for row in getattr(doc, "level_details", []):

        # Row-type detection:
        # - header row  -> has a Task link set
        # - average row -> no task, remark contains "Average"
        # - data row    -> everything else
        if row.task:
            row_type = "header"
        elif row.remark and "average" in row.remark.lower():
            row_type = "average"
        else:
            row_type = "data"

        cell_sr = ws_level.cell(row=row_num, column=1, value=sr_no)
        cell_sr.border = thin_border
        cell_sr.alignment = center_align

        if row_type == "header":
            cell_task = ws_level.cell(row=row_num, column=2, value=row.task or "")
            cell_task.font = bold_font
            cell_task.border = thin_border
            cell_task.alignment = left_align

            cell_subj = ws_level.cell(row=row_num, column=3, value=row.task_subject or "")
            cell_subj.font = bold_font
            cell_subj.border = thin_border
            cell_subj.alignment = left_align

            for c in range(4, 12):
                cell = ws_level.cell(row=row_num, column=c, value="")
                cell.border = thin_border
                cell.alignment = center_align

        elif row_type == "average":
            for c in range(2, 10):
                cell = ws_level.cell(row=row_num, column=c, value="")
                cell.border = thin_border
                cell.alignment = center_align

            avg_cell = ws_level.cell(row=row_num, column=10, value=flt(row.average_rl))
            avg_cell.font = bold_font
            avg_cell.border = thin_border
            avg_cell.alignment = center_align
            avg_cell.number_format = '0.000'

            remark_cell = ws_level.cell(row=row_num, column=11, value=row.remark or "")
            remark_cell.font = bold_font
            remark_cell.border = thin_border
            remark_cell.alignment = left_align

        else:  # data row
            for c in (2, 3):
                cell = ws_level.cell(row=row_num, column=c, value="")
                cell.border = thin_border
                cell.alignment = center_align

            data_fields = [
                ("design", 4), ("bs", 5), ("is", 6),
                ("fs", 7), ("hi", 8), ("rl", 9)
            ]
            for fieldname, col in data_fields:
                val = flt(row.get(fieldname))
                cell = ws_level.cell(row=row_num, column=col, value=val)
                cell.border = thin_border
                cell.alignment = center_align
                cell.number_format = '0.000'

            avg_cell = ws_level.cell(row=row_num, column=10, value="")
            avg_cell.border = thin_border
            avg_cell.alignment = center_align

            remark_cell = ws_level.cell(row=row_num, column=11, value=row.remark or "")
            remark_cell.border = thin_border
            remark_cell.alignment = left_align

        sr_no += 1
        row_num += 1

    for col in ws_level.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws_level.column_dimensions[column].width = (max_length + 2)
    ws_level.column_dimensions['A'].width = 8 

  
    # --- Level Data Sheet: EXACT copy of the on-screen grid ---
    matrix_json = doc.get("level_data_json")
    matrix = None
    if matrix_json:
        try:
            matrix = json.loads(matrix_json)
        except Exception:
            matrix = None
    if matrix and matrix.get("columns"):
        ws_leveldata = wb.create_sheet("Level Data")

        columns = matrix["columns"]
        rows = matrix.get("rows", [])

        total_cols = 2 + len(columns)  
        last_leveldata_col = get_column_letter(total_cols)

        ws_leveldata.merge_cells(f'A1:{last_leveldata_col}1')
        ws_leveldata['A1'] = company_name
        ws_leveldata['A1'].font = Font(bold=True, size=14)
        ws_leveldata['A1'].alignment = center_align

        ws_leveldata.merge_cells(f'A2:{last_leveldata_col}2')
        ws_leveldata['A2'] = project_doc.project_name or project_doc.name
        ws_leveldata['A2'].font = Font(bold=True, size=12)
        ws_leveldata['A2'].alignment = center_align

        ws_leveldata.merge_cells(f'A3:{last_leveldata_col}3')
        ws_leveldata['A3'] = f"RA Bill Number: {doc.name}"
        ws_leveldata['A3'].font = Font(bold=True, size=12)
        ws_leveldata['A3'].alignment = center_align

        ws_leveldata.merge_cells(f'A4:{last_leveldata_col}4')
        ws_leveldata['A4'] = "LEVEL DATA"
        ws_leveldata['A4'].font = Font(bold=True, size=12)
        ws_leveldata['A4'].alignment = center_align

        ws_leveldata.append([])   # blank row -> row 5

        # Header row — exactly "Sr.", "Particular", then each grid column, in order
        headers = ["Sr.No", "Particular"] + columns
        ws_leveldata.append(headers)   # -> row 6

        for col_num in range(1, total_cols + 1):
            cell = ws_leveldata.cell(row=6, column=col_num)
            cell.font = bold_font
            cell.alignment = center_align if col_num != 2 else left_align
            cell.border = thin_border

        # Data rows — exactly as they appear in the grid, no grouping/merging
        for idx, row in enumerate(rows, start=1):
            r = idx + 6  # row 7 onward

            cell_sr = ws_leveldata.cell(row=r, column=1, value=idx)
            cell_sr.border = thin_border
            cell_sr.alignment = center_align

            cell_particular = ws_leveldata.cell(row=r, column=2, value=row.get("particular", ""))
            cell_particular.font = bold_font
            cell_particular.border = thin_border
            cell_particular.alignment = left_align

            values = row.get("values") or {}

            for c_idx, col in enumerate(columns):
                raw_val = values.get(col, "")

                # copy as-is: numeric string -> number, otherwise leave as text
                val = raw_val
                if raw_val not in (None, ""):
                    try:
                        val = flt(raw_val)
                    except Exception:
                        val = raw_val

                cell = ws_leveldata.cell(row=r, column=3 + c_idx, value=val)
                cell.border = thin_border
                cell.alignment = center_align

        for col in ws_leveldata.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws_leveldata.column_dimensions[column].width = (max_length + 2)
        ws_leveldata.column_dimensions['A'].width = 8

    wb.move_sheet(summary_ws.title, offset=-len(wb.sheetnames))

    file_data = BytesIO()
    wb.save(file_data)

    frappe.response['filename'] = f"RA_Bill_{doc.name}.xlsx"
    frappe.response['filecontent'] = file_data.getvalue()
    frappe.response['type'] = 'binary'


# ----
@frappe.whitelist()
def get_steel_details(project, from_date, to_date):
    if not project or not from_date or not to_date:
        return []

    query = """
        SELECT 
            sed.item_code as item,
            sed.custom_task as task,
            sed.custom_subtask as subtask,
            sed.custom_diameter_of_steel as diamter_of_bar,
            sed.uom as unit,
            sed.qty as qty,
            se.name
        FROM 
            `tabStock Entry` se
        JOIN 
            `tabStock Entry Detail` sed ON se.name = sed.parent
        JOIN
            `tabItem` item ON sed.item_code = item.name
        WHERE 
            (se.project = %s OR sed.project = %s)
            AND se.posting_date >= %s
            AND se.stock_entry_type = 'Material Issue'
            AND se.posting_date <= %s
            AND se.docstatus = 1
            AND LOWER(item.custom_item_type) = 'steel'
    """
    
    data = frappe.db.sql(query, (project, project, from_date, to_date), as_dict=1)
    return data

@frappe.whitelist()
def get_level_sheet_details(project):
    if not project:
        return []

    result = []

    sheets = frappe.get_all(
        "Task Level Sheet",
        filters={
            "project": project,
            "docstatus": 1
        },
        fields=[
            "name",
            "average"
        ],
        order_by="creation asc"
    )

    hierarchy_fields = [
        "task_level10",
        "task_level9",
        "task_level8",
        "task_level7",
        "task_level6",
        "task_level5",
        "task_level4",
        "task_level3",
        "task_level2",
        "task_level1",
        "task"
    ]

    for sheet in sheets:

        sheet_doc = frappe.get_doc("Task Level Sheet", sheet.name)

        rows = sheet_doc.get("level_sheet_details") or []

        last_task = None

        for field in hierarchy_fields:
            if sheet_doc.get(field):
                last_task = sheet_doc.get(field)
                break

        last_task_subject = ""

        if last_task:
            last_task_subject = (
                frappe.db.get_value("Task", last_task, "subject") or ""
            )

        result.append({
            "is_header": 1,
            "task": last_task,
            "task_subject": last_task_subject
        })

        for d in rows:
            result.append({
                "design": d.get("design"),
                "bs": d.get("bs"),
                "is": d.get("is"),
                "fs": d.get("fs"),
                "hi": d.get("hi"),
                "rl": d.get("rl"),
                "remark": d.get("remark")
            })

        result.append({
            "is_average": 1,
            "average_rl": sheet.average,
            "remark": f"{last_task_subject} Average"
        })

    return result

@frappe.whitelist()
def get_level_matrix(project):

    if not project:
        return {}

    project_doc = frappe.get_doc("Project", project)

    columns = []

    for d in project_doc.custom_data_sheet_column:
        if d.parameter:
            columns.append(d.parameter)

    tasks = frappe.get_all(
        "Task",
        filters={
            "project": project,
            "custom_is_level_task": 1
        },
        fields=[
            "name",
            "subject",
            "parent_task",
            "custom_average_level"
        ]
    )

    rows = {}

    for task in tasks:

        parent_subject = frappe.db.get_value(
            "Task",
            task.parent_task,
            "subject"
        ) or ""

        if parent_subject not in rows:

            rows[parent_subject] = {
                "particular": parent_subject,
                "task_id": task.parent_task,
                "values": {}
            }

            for col in columns:
                rows[parent_subject]["values"][col] = ""

        rows[parent_subject]["values"][task.subject] = task.custom_average_level

    return {
        "columns": columns,
        "rows": list(rows.values())
    }

@frappe.whitelist()
def calculate_level_matrix(project, matrix):
    """
    Recalculate the Level Matrix grid using the Project's
    custom_data_sheet_column (Parameter -> Abbr) and
    custom_data_sheet_formulas (Parameter -> Formula) child tables.

    `matrix` is the same {columns, rows: [{particular, values: {parameter: value}}]}
    shape produced by get_level_matrix / rendered by render_level_matrix in JS,
    with whatever the user has manually typed into the grid.
    """

    if not project:
        frappe.throw(_("Project is required"))

    matrix = frappe.parse_json(matrix)
    project_doc = frappe.get_doc("Project", project)

    column_map = {}
    abbr_map = {}

    for d in project_doc.custom_data_sheet_column:
        if not d.parameter:
            continue
        column_map[d.parameter] = d.abbr
        if d.abbr:
            abbr_map[d.abbr] = d.parameter

    formula_map = {}
    for d in project_doc.custom_data_sheet_formulas:
        if d.parameter and d.formula:
            formula_map[d.parameter] = d.formula.strip()

    sorted_abbrs = sorted(abbr_map.keys(), key=len, reverse=True)

    for row in matrix.get("rows", []):
        values = row.get("values") or {}

        context = {}
        for parameter, value in values.items():
            if parameter in formula_map:
                continue
            abbr = column_map.get(parameter)
            if not abbr:
                continue
            if value in (None, ""):
                continue
            try:
                context[abbr] = flt(value)
            except Exception:
                pass

        pending = dict(formula_map)

        while pending:
            resolved_this_pass = []

            for parameter, formula in pending.items():
                expr = formula
                unresolved = False

                for abbr in sorted_abbrs:
                    pattern = r"\b{}\b".format(re.escape(abbr))
                    if re.search(pattern, expr):
                        if abbr in context:
                            expr = re.sub(pattern, "({})".format(context[abbr]), expr)
                        else:
                            unresolved = True

                if unresolved:
                    continue

                try:
                    result = flt(eval(expr, {"__builtins__": {}}, {}))
                except Exception:
                    result = 0

                values[parameter] = round(result, 3)

                abbr = column_map.get(parameter)
                if abbr:
                    context[abbr] = result

                resolved_this_pass.append(parameter)

            for parameter in resolved_this_pass:
                pending.pop(parameter, None)
\
            if not resolved_this_pass:
                break

        row["values"] = values

    return matrix

@frappe.whitelist()
def download_steel_template(rows):
    rows = frappe.parse_json(rows)

    wb = Workbook()
    ws = wb.active
    ws.title = "Steel Template"

    headers = [
        "Stage",
        "Stage Subject",
        "Task",
        "Task Subject",
        "Task Level1",
        "Task Level1 Subject",
        "Task Level2",
        "Task Level2 Subject",
        "Task Level3",
        "Task Level3 Subject",
        "Task Level4",
        "Task Level4 Subject",
        "Task Level5",
        "Task Level5 Subject",
        "Task Level6",
        "Task Level6 Subject",
        "Task Level7",
        "Task Level7 Subject",
        "Task Level8",
        "Task Level8 Subject",
        "Task Level9",
        "Task Level9 Subject",
        "Task Level10",
        "Task Level10 Subject",
        "Location Of The Bar",
        "Nos",
        "Length",
        "Width",
        "Depth",
        "Ftg. Depth",
        "Dia Of Bar",
        "Spacing",
        "Bar Nos",
        "Bar Length",
        "Column Height",
        "Top Beam Depth",
        "8 MM Reinforcement",
        "10 MM Reinforcement",
        "12 MM Reinforcement",
        "16 MM Reinforcement",
        "20 MM Reinforcement",
        "25 MM Reinforcement",
        "28 MM Reinforcement",
        "32 MM Reinforcement",
        "Remark"
    ]

    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for d in rows:
        ws.append([
            d.get("stage"),
            d.get("stage_subject"),
            d.get("task"),
            d.get("task_subject"),
            d.get("task_level1"),
            d.get("level1_subject"),
            d.get("task_level2"),
            d.get("level2_subject"),
            d.get("task_level3"),
            d.get("level3_subject"),
            d.get("task_level4"),
            d.get("level4_subject"),
            d.get("task_level5"),
            d.get("level5_subject"),
            d.get("task_level6"),
            d.get("level6_subject"),
            d.get("task_level7"),
            d.get("level7_subject"),
            d.get("task_level8"),
            d.get("level8_subject"),
            d.get("task_level9"),
            d.get("level9_subject"),
            d.get("task_level10"),
            d.get("level10_subject"),
            "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""," ",
        ])
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column].width = max_length + 2
    ws.column_dimensions['A'].width =8
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.response["filename"] = "Steel_Template.xlsx"
    frappe.response["filecontent"] = output.getvalue()
    frappe.response["type"] = "download"


LABEL_TO_FIELD = {
    "Stage": "stage",
    "Stage Subject": "stage_subject",
    "Task": "task",
    "Task Subject": "task_subject",
    "Task Level1": "task_level1", "Task Level1 Subject": "level1_subject",
    "Task Level2": "task_level2", "Task Level2 Subject": "level2_subject",
    "Task Level3": "task_level3", "Task Level3 Subject": "level3_subject",
    "Task Level4": "task_level4", "Task Level4 Subject": "level4_subject",
    "Task Level5": "task_level5", "Task Level5 Subject": "level5_subject",
    "Task Level6": "task_level6", "Task Level6 Subject": "level6_subject",
    "Task Level7": "task_level7", "Task Level7 Subject": "level7_subject",
    "Task Level8": "task_level8", "Task Level8 Subject": "level8_subject",
    "Task Level9": "task_level9", "Task Level9 Subject": "level9_subject",
    "Task Level10": "task_level10", "Task Level10 Subject": "level10_subject",
    "Location Of The Bar": "location_of_the_bar",
    "Nos": "nos",
    "Length": "length",
    "Width": "width",
    "Depth": "depth",
    "Ftg. Depth": "ftg_depth",
    "Dia Of Bar": "dia_of_bar",
    "Spacing": "spacing",
    "Bar Nos": "bar_nos",
    "Bar Length": "bar_length",
    "Column Height": "column_height",
    "Top Beam Depth": "top_beam_depth",
    "8 MM Reinforcement": "8_mm_reinforcement",
    "10 MM Reinforcement": "10_mm_reinforcement",
    "12 MM Reinforcement": "12_mm_reinforcement",
    "16 MM Reinforcement": "16_mm_reinforcement",
    "20 MM Reinforcement": "20_mm_reinforcement",
    "25 MM Reinforcement": "25_mm_reinforcement",
    "28 MM Reinforcement": "28_mm_reinforcement",
    "32 MM Reinforcement": "32_mm_reinforcement",
    "Remark": "remark",
}

EDITABLE_FIELDS = [
    "location_of_the_bar","nos", "length", "width", "depth", "ftg_depth",
    "dia_of_bar", "spacing", "bar_nos" ,"bar_length", "column_height",
    "top_beam_depth",
    "8_mm_reinforcement", "10_mm_reinforcement", "12_mm_reinforcement",
    "16_mm_reinforcement", "20_mm_reinforcement", "25_mm_reinforcement",
    "28_mm_reinforcement", "32_mm_reinforcement",
    "remark",
]

HIERARCHY_FIELDS = [f"task_level{i}" for i in range(1, 11)]


def _row_key(row_like):

    parts = [row_like.get("stage") if isinstance(row_like, dict) else row_like.stage,
             row_like.get("task") if isinstance(row_like, dict) else row_like.task]

    for f in HIERARCHY_FIELDS:
        val = row_like.get(f) if isinstance(row_like, dict) else row_like.get(f)
        if val:
            parts.append(val)

    return tuple(parts)


@frappe.whitelist()
def import_steel_template(docname, file_url):

    doc = frappe.get_doc("RA Billing", docname)

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_path = file_doc.get_full_path()

    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_index = {}
    for idx, label in enumerate(header_row):
        if label and str(label).strip() in LABEL_TO_FIELD:
            col_index[LABEL_TO_FIELD[str(label).strip()]] = idx

    existing_by_key = {}
    for row in doc.ra_steel_details:
        existing_by_key[_row_key(row)] = row

    updated = 0
    unmatched_rows = []

    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if excel_row is None or all(v in (None, "") for v in excel_row):
            continue  # skip fully blank rows

        row_data = {}
        for field, idx in col_index.items():
            if idx < len(excel_row):
                row_data[field] = excel_row[idx]

        key = _row_key(row_data)
        child = existing_by_key.get(key)

        if not child:
            label = row_data.get("level1_subject") or row_data.get("task_subject") or str(key)
            unmatched_rows.append(label)
            continue

        for field in EDITABLE_FIELDS:
            value = row_data.get(field)
            if value not in (None, ""):
                child.set(field, value)

    doc.save()

    return {
        "updated": updated,
        "unmatched": len(unmatched_rows),
        "unmatched_rows": unmatched_rows,
    }

def get_previous_stage_totals(project, current_name):
    """
    Returns {stage: {"total_bill_quantity": x, "total_bill_amount": y}}
    from the most recent SUBMITTED RA Billing for this project
    (excluding the current document).
    """
    if not project:
        return {}

    previous_name = frappe.db.get_value(
        "RA Billing",
        {
            "project": project,
            "docstatus": 1,
            "name": ["!=", current_name or ""]
        },
        "name",
        order_by="creation desc"
    )

    if not previous_name:
        return {}

    previous_doc = frappe.get_doc("RA Billing", previous_name)

    totals = {}
    for row in previous_doc.ra_abstract_details:
        totals[row.stage] = {
            "total_bill_quantity": flt(row.total_bill_quantity),
            "total_bill_amount": flt(row.total_bill_amount),
        }

    return totals

def _safe_sheet_name(name, used_names):
    """Excel sheet names: max 31 chars, no : \\ / ? * [ ], must be unique in workbook."""
    for ch in ['\\', '/', '?', '*', '[', ']', ':']:
        name = name.replace(ch, '-')
    name = name[:31]

    base = name
    counter = 1
    while name in used_names:
        suffix = f"_{counter}"
        name = base[: 31 - len(suffix)] + suffix
        counter += 1

    used_names.add(name)
    return name

def build_summary_sheet(wb, doc, used_sheet_names):
    """
    Builds the 'Summary' sheet for a single RA Billing document:
    one row showing the Project name and total amount, followed by a
    Total row, then (if with_tax) Tax Category rows with their rate,
    and finally the Grand Total row.
    """
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    with_tax = bool(doc.get("with_tax"))

    ws = wb.create_sheet(_safe_sheet_name("Summary", used_sheet_names))

    if with_tax:
        headers = ["Sr. No", "Description", "Tax Rate", "Amount Rs."]
        rate_col = 3
        amt_col = 4
    else:
        headers = ["Sr. No", "Description", "Amount Rs."]
        rate_col = None
        amt_col = 3

    total_cols = len(headers)
    last_col = get_column_letter(total_cols)

    project_doc = frappe.get_doc("Project", doc.project)
    company_name = project_doc.company

    # ---- 4-row header block (back to original) ----
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = company_name
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_align

    ws.merge_cells(f'A2:{last_col}2')
    ws['A2'] = project_doc.project_name or project_doc.name
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = center_align

    ws.merge_cells(f'A3:{last_col}3')
    ws['A3'] = f"R.A. Bill No-{doc.name}"
    ws['A3'].font = Font(bold=True, size=13)
    ws['A3'].alignment = center_align

    ws.merge_cells(f'A4:{last_col}4')
    ws['A4'] = "Total Summary"
    ws['A4'].font = Font(bold=True, size=12)
    ws['A4'].alignment = center_align

    ws.append([])          # blank row -> row 5
    ws.append(headers)     # header row -> row 6

    for col_num in range(1, total_cols + 1):
        cell = ws.cell(row=6, column=col_num)
        cell.font = bold_font
        cell.alignment = center_align if col_num != 2 else left_align
        cell.border = thin_border

    row_num = 7

    # ---- Single row: Description = Project name, Amount = grand_total ----
    project_display_name = project_doc.project_name or project_doc.name
    grand_total = flt(doc.get("grand_total"))

    cell_sr = ws.cell(row=row_num, column=1, value=1)
    cell_sr.border = thin_border
    cell_sr.alignment = left_align

    cell_desc = ws.cell(row=row_num, column=2, value=project_display_name)
    cell_desc.border = thin_border
    cell_desc.alignment = left_align

    if rate_col:
        ws.cell(row=row_num, column=rate_col, value="").border = thin_border

    amt_cell = ws.cell(row=row_num, column=amt_col, value=grand_total)
    amt_cell.border = thin_border
    amt_cell.alignment = center_align
    amt_cell.number_format = '0.00'

    row_num += 1

    # ---- Total row ----
    cell_label = ws.cell(row=row_num, column=2, value="Total")
    cell_label.font = bold_font
    cell_label.border = thin_border
    ws.cell(row=row_num, column=1, value="").border = thin_border
    if rate_col:
        ws.cell(row=row_num, column=rate_col, value="").border = thin_border

    total_cell = ws.cell(row=row_num, column=amt_col, value=grand_total)
    total_cell.font = bold_font
    total_cell.border = thin_border
    total_cell.number_format = '0.00'
    row_num += 1

    # ---- Tax rows (only if With Tax) ----
    tax_total = 0.0
    if with_tax:
        for tax_row in doc.get("tax_details", []):
            ws.cell(row=row_num, column=1, value="").border = thin_border
            ws.cell(row=row_num, column=2, value=tax_row.tax_category or "").border = thin_border

            rate_cell = ws.cell(row=row_num, column=rate_col, value=flt(tax_row.tax_rate))
            rate_cell.border = thin_border
            rate_cell.alignment = center_align
            rate_cell.number_format = '0.00"%"'

            tax_amt_cell = ws.cell(row=row_num, column=amt_col, value=flt(tax_row.tax_amount))
            tax_amt_cell.border = thin_border
            tax_amt_cell.number_format = '0.00'

            tax_total += flt(tax_row.tax_amount)
            row_num += 1

    # ---- Grand Total row ----
    cell_label = ws.cell(row=row_num, column=2, value="Grand Total")
    cell_label.font = bold_font
    cell_label.border = thin_border
    ws.cell(row=row_num, column=1, value="").border = thin_border
    if rate_col:
        ws.cell(row=row_num, column=rate_col, value="").border = thin_border

    final_total = flt(doc.get("final_grand_total")) or (grand_total + tax_total)
    gt_cell = ws.cell(row=row_num, column=amt_col, value=final_total)
    gt_cell.font = bold_font
    gt_cell.border = thin_border
    gt_cell.number_format = '0.00'

    for c in range(1, total_cols + 1):
        ws.cell(row=row_num, column=c).fill = openpyxl.styles.PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )

    row_num += 1

    row_num += 2
    generated_on = frappe.utils.format_datetime(
        frappe.utils.now_datetime(), "dd-MM-yyyy HH:mm:ss"
    )
    gen_cell = ws.cell(row=row_num, column=2, value=f"Generated On: {generated_on}")
    gen_cell.font = Font(italic=True, size=10)
    gen_cell.alignment = left_align

    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column].width = (max_length + 2)
    ws.column_dimensions['A'].width = 8

    return ws
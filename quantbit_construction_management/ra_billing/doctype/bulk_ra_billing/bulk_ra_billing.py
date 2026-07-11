# Copyright (c) 2026, QTPL and contributors
# For license information, please see license.txt

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.utils import flt
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import math
import json

class BulkRABilling(Document):
	pass

@frappe.whitelist()
def get_projects_for_site(site):
    if not site:
        return []

    projects = frappe.get_all(
        "Project",
        filters={"custom_site": site},
        fields=["name", "project_name"]
    )

    return projects

def _safe_sheet_name(name, used_names):
    """Excel sheet names: max 31 chars, no : \\ / ? * [ ], must be unique in workbook."""
    for ch in ['\\', '/', '?', '*', '[', ']', ':']:
        name = name.replace(ch, '-')
    name = name[:31]

    base = name
    counter = 1
    while name in used_names:
        suffix = f"_{counter}"
        name = base[: 31 - len(suffix)] + suffix
        counter += 1

    used_names.add(name)
    return name


def build_ra_sheets_into_workbook(wb, ra_billing_name, sheet_prefix, used_sheet_names):
    """
    Builds Abstract / Measurement / Steel / Level Details / Level Data sheets
    for ONE RA Billing document into the given workbook `wb`, with sheet
    names prefixed by `sheet_prefix` (kept unique across the whole workbook
    via `used_sheet_names`).

    Returns: {"project_name": str, "abstract_total": float}
    """
    doc = frappe.get_doc("RA Billing", ra_billing_name)
    project_doc = frappe.get_doc("Project", doc.project)
    company_name = project_doc.company
    display_project_name = project_doc.project_name or project_doc.name

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    task_desc_cache = {}
    steel_task_totals = {}
    steel_stage_totals = {}
    steel_stage_weights = {}

    def get_task_desc(task_id):
        if not task_id:
            return ""
        if task_id not in task_desc_cache:
            desc = frappe.db.get_value("Task", task_id, "description")
            task_desc_cache[task_id] = frappe.utils.strip_html(desc) if desc else ""
        return task_desc_cache[task_id]

    BAR_DIAMETERS = [8, 10, 12, 16, 20, 25, 28, 32]
    BAR_WEIGHT_PER_METER = {
        d: math.trunc((d ** 2) / 162.0 * 1000) / 1000 for d in BAR_DIAMETERS
    }

    def get_row_reinforcement(row):
        values = {}
        for d in BAR_DIAMETERS:
            val = flt(row.get(f"{d}_mm_reinforcement"))
            if val:
                values[d] = val
        return values

    def get_row_weight(row):
        total = 0.0
        for d, length in get_row_reinforcement(row).items():
            total += length * BAR_WEIGHT_PER_METER[d]
        return total

    LEVEL_FIELDS = [f"task_level{i}" for i in range(1, 10)]
    LEVEL_SUBJECT_FIELDS = [f"level{i}_subject" for i in range(1, 10)]

    def get_row_hierarchy(row):
        path = []
        if row.get("stage"):
            path.append((row.get("stage"), row.get("stage_subject") or row.get("stage")))
        if row.get("task"):
            path.append((row.get("task"), row.get("task_subject") or row.get("task")))
        for lvl_field, subj_field in zip(LEVEL_FIELDS, LEVEL_SUBJECT_FIELDS):
            if row.get(lvl_field):
                path.append((row.get(lvl_field), row.get(subj_field) or row.get(lvl_field)))
        return path

    abstract_total = 0.0

 # ============ ABSTRACT SHEET ============
    ws_abstract = wb.create_sheet(_safe_sheet_name(f"{sheet_prefix} Abstract", used_sheet_names))

    abs_headers = [
        "Sr. No", "Stage", "Description", "UOM", "Rate",
        "Previous Bill Qty", "This Bill Qty", "Total Bill Qty",
        "Previous Bill Amount", "This Bill Amount", "Total Bill Amount",
        "Remarks"
    ]
    abs_total_cols = len(abs_headers)
    abs_last_col = get_column_letter(abs_total_cols)

    ws_abstract.merge_cells(f'A1:{abs_last_col}1')
    ws_abstract['A1'] = company_name
    ws_abstract['A1'].font = Font(bold=True, size=14)
    ws_abstract['A1'].alignment = center_align

    ws_abstract.merge_cells(f'A2:{abs_last_col}2')
    ws_abstract['A2'] = display_project_name
    ws_abstract['A2'].font = Font(bold=True, size=12)
    ws_abstract['A2'].alignment = center_align

    ws_abstract.merge_cells(f'A3:{abs_last_col}3')
    ws_abstract['A3'] = f"RA Bill Number: {doc.name}"
    ws_abstract['A3'].font = Font(bold=True, size=12)
    ws_abstract['A3'].alignment = center_align

    ws_abstract.merge_cells(f'A4:{abs_last_col}4')
    ws_abstract['A4'] = "ABSTRACT SHEET"
    ws_abstract['A4'].font = Font(bold=True, size=12)
    ws_abstract['A4'].alignment = center_align

    ws_abstract.append([])            # blank row -> row 5
    ws_abstract.append(abs_headers)   # header row -> row 6

    for col_num in range(1, abs_total_cols + 1):
        cell = ws_abstract.cell(row=6, column=col_num)
        cell.font = bold_font
        cell.alignment = center_align if col_num not in (2, 3, 12) else left_align
        cell.border = thin_border

    row_num = 7
    sr_no = 1
    grand_this_amt = 0.0
    grand_total_amt = 0.0

    for row in getattr(doc, "ra_abstract_details", []):
        rate = flt(row.rate)
        prev_qty = flt(row.previous_bill_quantity)
        this_qty = flt(row.billed_quantity)
        total_qty = flt(row.total_bill_quantity) or (prev_qty + this_qty)

        prev_amt = flt(row.previous_bill_amount)
        this_amt = flt(row.amount)
        total_amt = flt(row.total_bill_amount) or (prev_amt + this_amt)

        cell_sr = ws_abstract.cell(row=row_num, column=1, value=sr_no)
        cell_sr.border = thin_border
        cell_sr.alignment = left_align

        cell_stage = ws_abstract.cell(row=row_num, column=2, value=row.stage_subject or "")
        cell_stage.font = bold_font
        cell_stage.border = thin_border
        cell_stage.alignment = left_align

        cell_desc = ws_abstract.cell(row=row_num, column=3, value=row.description or "")
        cell_desc.border = thin_border
        cell_desc.alignment = left_align

        cell_uom = ws_abstract.cell(row=row_num, column=4, value=row.uom or "")
        cell_uom.border = thin_border
        cell_uom.alignment = center_align

        rate_cell = ws_abstract.cell(row=row_num, column=5, value=rate)
        rate_cell.border = thin_border
        rate_cell.alignment = center_align
        rate_cell.number_format = '0.00'

        prev_qty_cell = ws_abstract.cell(row=row_num, column=6, value=prev_qty)
        prev_qty_cell.border = thin_border
        prev_qty_cell.alignment = center_align
        prev_qty_cell.number_format = '0.00'

        this_qty_cell = ws_abstract.cell(row=row_num, column=7, value=this_qty)
        this_qty_cell.border = thin_border
        this_qty_cell.alignment = center_align
        this_qty_cell.number_format = '0.00'

        total_qty_cell = ws_abstract.cell(row=row_num, column=8, value=total_qty)
        total_qty_cell.border = thin_border
        total_qty_cell.alignment = center_align
        total_qty_cell.number_format = '0.00'

        prev_amt_cell = ws_abstract.cell(row=row_num, column=9, value=prev_amt)
        prev_amt_cell.border = thin_border
        prev_amt_cell.alignment = center_align
        prev_amt_cell.number_format = '0.00'

        this_amt_cell = ws_abstract.cell(row=row_num, column=10, value=this_amt)
        this_amt_cell.border = thin_border
        this_amt_cell.alignment = center_align
        this_amt_cell.number_format = '0.00'

        total_amt_cell = ws_abstract.cell(row=row_num, column=11, value=total_amt)
        total_amt_cell.border = thin_border
        total_amt_cell.alignment = center_align
        total_amt_cell.number_format = '0.00'

        remarks_cell = ws_abstract.cell(row=row_num, column=12, value=getattr(row, "remarks", "") or "")
        remarks_cell.border = thin_border
        remarks_cell.alignment = left_align

        grand_this_amt += this_amt
        grand_total_amt += total_amt
        abstract_total += this_amt

        sr_no += 1
        row_num += 1

    # Grand Total row
    cell_sr = ws_abstract.cell(row=row_num, column=1, value="")
    cell_sr.border = thin_border

    cell_desc = ws_abstract.cell(row=row_num, column=2, value="GRAND TOTAL")
    cell_desc.font = bold_font
    cell_desc.border = thin_border
    cell_desc.alignment = left_align

    for c in (3, 4, 5, 6, 7, 8, 9):
        cell = ws_abstract.cell(row=row_num, column=c, value="")
        cell.border = thin_border

    gt_this_amt = ws_abstract.cell(row=row_num, column=10, value=grand_this_amt)
    gt_this_amt.font = bold_font
    gt_this_amt.border = thin_border
    gt_this_amt.alignment = center_align
    gt_this_amt.number_format = '0.00'

    gt_total_amt = ws_abstract.cell(row=row_num, column=11, value=grand_total_amt)
    gt_total_amt.font = bold_font
    gt_total_amt.border = thin_border
    gt_total_amt.alignment = center_align
    gt_total_amt.number_format = '0.00'

    cell_remark = ws_abstract.cell(row=row_num, column=12, value="")
    cell_remark.border = thin_border

    row_num += 1

    for col in ws_abstract.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws_abstract.column_dimensions[column].width = (max_length + 2)
    # ============ MEASUREMENT SHEET ============
    ws_meas = wb.create_sheet(_safe_sheet_name(f"{sheet_prefix} Measurement", used_sheet_names))

    ws_meas.merge_cells('A1:J1')
    ws_meas['A1'] = company_name
    ws_meas['A1'].font = Font(bold=True, size=14)
    ws_meas['A1'].alignment = center_align

    ws_meas.merge_cells('A2:J2')
    ws_meas['A2'] = display_project_name
    ws_meas['A2'].font = Font(bold=True, size=12)
    ws_meas['A2'].alignment = center_align

    ws_meas.merge_cells('A3:J3')
    ws_meas['A3'] = f"RA Bill Number: {doc.name}"
    ws_meas['A3'].font = Font(bold=True, size=12)
    ws_meas['A3'].alignment = center_align

    ws_meas.merge_cells('A4:J4')
    ws_meas['A4'] = "MEASUREMENT SHEET"
    ws_meas['A4'].font = Font(bold=True, size=12)
    ws_meas['A4'].alignment = center_align

    ws_meas.append([])
    meas_headers = ["Sr. No", "Content", "Description", "No.1", "No.2", "Length", "Width", "Depth", "Unit", "Qty"]
    ws_meas.append(meas_headers)

    for col_num in range(1, 11):
        cell = ws_meas.cell(row=6, column=col_num)
        cell.font = bold_font
        cell.alignment = center_align if col_num not in (2, 3) else left_align
        cell.border = thin_border

    def get_or_create_node(container, order_list, node_id, subject):
        if node_id not in container:
            container[node_id] = {"subject": subject, "children": {}, "children_order": [], "rows": []}
            order_list.append(node_id)
        return container[node_id]

    # steel weight lookup keyed by (task, deepest_level_id)
    steel_subtask_weights = {}
    for steel_row in getattr(doc, "ra_steel_details", []):
        path = get_row_hierarchy(steel_row)
        deepest_id = path[-1][0] if path else None
        key = (steel_row.get("task"), deepest_id)
        steel_subtask_weights[key] = steel_subtask_weights.get(key, 0) + get_row_weight(steel_row)

    # which stages have at least one Kg-uom row (drives the Kg/MT stage summary rows)
    stage_has_kg = {}
    for row in getattr(doc, "ra_billing_details", []):
        stage_subj = row.stage_subject or "No Stage"
        if (row.uom or "").lower() in ["kg", "kilogram"]:
            stage_has_kg[stage_subj] = True

    def get_deepest_billing_task_id(row):
        path = get_row_hierarchy(row)
        return path[-1][0] if path else row.get("task")

    # build the full nested tree: stage -> task -> task_level1 -> ... -> task_level9
    tree_meas = {}
    node_order_meas = []

    for row in getattr(doc, "ra_billing_details", []):
        path = get_row_hierarchy(row)
        if not path:
            continue

        node = get_or_create_node(tree_meas, node_order_meas, path[0][0], path[0][1])
        for node_id, subject in path[1:]:
            node = get_or_create_node(node["children"], node["children_order"], node_id, subject)

        node["rows"].append(row)

    row_num = 7
    sr_no = 1

    def write_meas_leaf_row(r, row_label, indent, is_first_row):
        nonlocal row_num

        desc_val = (indent + row_label) if is_first_row else ""
        cell_desc = ws_meas.cell(row=row_num, column=2, value=desc_val)
        if is_first_row:
            cell_desc.font = bold_font
        cell_desc.border = thin_border
        cell_desc.alignment = left_align

        desc_text = r.description if getattr(r, "description", None) else get_task_desc(get_deepest_billing_task_id(r))
        cell_subdesc = ws_meas.cell(row=row_num, column=3, value=desc_text if is_first_row else "")
        cell_subdesc.border = thin_border
        cell_subdesc.alignment = left_align

        cell = ws_meas.cell(row=row_num, column=4, value=r.no1 or "")
        cell.border = thin_border; cell.alignment = center_align

        cell = ws_meas.cell(row=row_num, column=5, value=r.no2 or "")
        cell.border = thin_border; cell.alignment = center_align

        cell = ws_meas.cell(row=row_num, column=6, value=r.length or "")
        cell.border = thin_border; cell.alignment = center_align

        cell = ws_meas.cell(row=row_num, column=7, value=r.width or "")
        cell.border = thin_border; cell.alignment = center_align

        cell = ws_meas.cell(row=row_num, column=8, value=r.height or "")
        cell.border = thin_border; cell.alignment = center_align

        uom_val = r.uom or ""
        cell = ws_meas.cell(row=row_num, column=9, value=uom_val)
        cell.border = thin_border; cell.alignment = center_align

        qty = flt(r.quantity)
        if uom_val.lower() in ["kg", "kilogram"]:
            deepest_id = get_deepest_billing_task_id(r)
            display_qty = steel_subtask_weights.get((r.task, deepest_id), qty)
        else:
            display_qty = qty

        qty_cell = ws_meas.cell(row=row_num, column=10, value=display_qty)
        qty_cell.border = thin_border; qty_cell.alignment = center_align
        qty_cell.number_format = '0.00'

        row_num += 1

    def write_meas_node(node, depth):
        nonlocal row_num, sr_no

        indent = "  " * depth
        is_leaf = not node["children_order"]

        if is_leaf and node["rows"]:
            sr_row = row_num

            write_meas_leaf_row(node["rows"][0], node["subject"], indent, is_first_row=True)

            cell_sr = ws_meas.cell(row=sr_row, column=1, value=sr_no)
            cell_sr.border = thin_border; cell_sr.alignment = left_align
            sr_no += 1

            for r in node["rows"][1:]:
                write_meas_leaf_row(r, "", indent, is_first_row=False)

        else:
            cell_sr = ws_meas.cell(row=row_num, column=1, value=sr_no)
            cell_sr.border = thin_border; cell_sr.alignment = left_align
            sr_no += 1

            cell_desc = ws_meas.cell(row=row_num, column=2, value=indent + node["subject"])
            cell_desc.font = bold_font
            cell_desc.border = thin_border
            cell_desc.alignment = left_align

            for c in range(3, 11):
                cell = ws_meas.cell(row=row_num, column=c, value="")
                cell.border = thin_border

            row_num += 1

            for r in node["rows"]:
                write_meas_leaf_row(r, "", indent, is_first_row=False)

        for child_id in node["children_order"]:
            child = node["children"][child_id]
            write_meas_node(child, depth + 1)

    for stage_id in node_order_meas:
        stage_node = tree_meas[stage_id]

        write_meas_node(stage_node, 0)

        if stage_has_kg.get(stage_node["subject"]):
            stage_kg_total = steel_stage_weights.get(stage_node["subject"], 0)
            stage_mt_total = steel_stage_totals.get(stage_node["subject"], 0)

            cell_desc = ws_meas.cell(row=row_num, column=2, value="Stage Total (Kg)")
            cell_desc.font = bold_font
            cell_desc.border = thin_border
            cell_desc.alignment = left_align
            for c in [1] + list(range(3, 10)):
                cell = ws_meas.cell(row=row_num, column=c, value="")
                cell.border = thin_border
            kg_cell = ws_meas.cell(row=row_num, column=10, value=stage_kg_total)
            kg_cell.font = bold_font
            kg_cell.border = thin_border
            kg_cell.number_format = "0.00"
            row_num += 1

            cell_desc = ws_meas.cell(row=row_num, column=2, value="Stage Total (Metric Tonne)")
            cell_desc.font = bold_font
            cell_desc.border = thin_border
            cell_desc.alignment = left_align
            for c in [1] + list(range(3, 10)):
                cell = ws_meas.cell(row=row_num, column=c, value="")
                cell.border = thin_border
            mt_cell = ws_meas.cell(row=row_num, column=10, value=stage_mt_total)
            mt_cell.font = bold_font
            mt_cell.border = thin_border
            mt_cell.number_format = "0.000"
            row_num += 1

    for col in ws_meas.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws_meas.column_dimensions[column].width = (max_length + 2)

 # ============ STEEL SHEET ============
    ws_steel = wb.create_sheet(_safe_sheet_name(f"{sheet_prefix} Steel", used_sheet_names))

    steel_headers = (
        ["Sr.", "Description of Item", "Location Of The Bar", "Nos", "Length", "Width", "Depth",
         "Ftg. Depth", "Dia Of Bar", "Spacing", "Bar Nos", "Bar Length", "Column Height", "Top Beam Depth"]
        + [f"{d} MM" for d in BAR_DIAMETERS]
        + ["Remark"]
    )
    steel_total_cols = len(steel_headers)
    last_steel_col = get_column_letter(steel_total_cols)

    ws_steel.merge_cells(f'A1:{last_steel_col}1')
    ws_steel['A1'] = company_name
    ws_steel['A1'].font = Font(bold=True, size=14)
    ws_steel['A1'].alignment = center_align

    ws_steel.merge_cells(f'A2:{last_steel_col}2')
    ws_steel['A2'] = display_project_name
    ws_steel['A2'].font = Font(bold=True, size=12)
    ws_steel['A2'].alignment = center_align

    ws_steel.merge_cells(f'A3:{last_steel_col}3')
    ws_steel['A3'] = f"RA Bill Number: {doc.name}"
    ws_steel['A3'].font = Font(bold=True, size=12)
    ws_steel['A3'].alignment = center_align

    ws_steel.merge_cells(f'A4:{last_steel_col}4')
    ws_steel['A4'] = "STEEL DETAILS"
    ws_steel['A4'].font = Font(bold=True, size=12)
    ws_steel['A4'].alignment = center_align

    ws_steel.append([])
    ws_steel.append(steel_headers)

    for col_num in range(1, len(steel_headers) + 1):
        cell = ws_steel.cell(row=6, column=col_num)
        cell.font = bold_font
        cell.alignment = center_align if col_num != 2 else left_align
        cell.border = thin_border

    # --- build the nested tree from ra_steel_details ---
    tree = {}
    node_order = []

    for row in getattr(doc, "ra_steel_details", []):
        path = get_row_hierarchy(row)
        if not path:
            continue
        node = get_or_create_node(tree, node_order, path[0][0], path[0][1])
        for node_id, subject in path[1:]:
            node = get_or_create_node(node["children"], node["children_order"], node_id, subject)
        node["rows"].append(row)

    row_num = 7
    sr_no = 1
    grand_total_weight = 0.0
    diameter_length_totals = {d: 0.0 for d in BAR_DIAMETERS}

    def write_leaf_row(r, row_label, indent, is_first_row):
        nonlocal row_num

        desc_val = (indent + row_label) if is_first_row else ""
        cell_desc = ws_steel.cell(row=row_num, column=2, value=desc_val)
        if is_first_row:
            cell_desc.font = bold_font
        cell_desc.border = thin_border
        cell_desc.alignment = left_align

        loc_cell = ws_steel.cell(row=row_num, column=3, value=r.get("location_of_the_bar") or "")
        loc_cell.border = thin_border
        loc_cell.alignment = left_align

        plain_fields = [
            ("nos", 4), ("length", 5), ("width", 6), ("depth", 7),
            ("ftg_depth", 8), ("dia_of_bar", 9), ("spacing", 10),
            ("bar_nos", 11), ("bar_length", 12),
            ("column_height", 13), ("top_beam_depth", 14)
        ]
        for fieldname, col in plain_fields:
            val = flt(r.get(fieldname))
            cell = ws_steel.cell(row=row_num, column=col, value=val if val else "")
            cell.border = thin_border
            cell.alignment = center_align
            if val:
                cell.number_format = '0.00'

        reinforcement = get_row_reinforcement(r)
        for idx, d in enumerate(BAR_DIAMETERS):
            val = reinforcement.get(d, "")
            cell = ws_steel.cell(row=row_num, column=15 + idx, value=val)
            cell.border = thin_border
            cell.alignment = center_align
            if val:
                cell.number_format = '0.00'
                diameter_length_totals[d] += val

        remark_cell = ws_steel.cell(row=row_num, column=23, value=r.get("remark") or "")
        remark_cell.border = thin_border
        remark_cell.alignment = left_align

        row_num += 1
        return get_row_weight(r)

    def write_node(node, depth, task_key=None):
        nonlocal row_num, sr_no, grand_total_weight
        indent = "  " * depth
        node_weight = 0.0
        is_leaf = not node["children_order"]

        if is_leaf and node["rows"]:
            sr_row = row_num
            node_weight += write_leaf_row(node["rows"][0], node["subject"], indent, True)

            cell_sr = ws_steel.cell(row=sr_row, column=1, value=sr_no)
            cell_sr.border = thin_border
            cell_sr.alignment = left_align
            sr_no += 1

            for r in node["rows"][1:]:
                node_weight += write_leaf_row(r, node["subject"], indent, False)
        else:
            cell_sr = ws_steel.cell(row=row_num, column=1, value=sr_no)
            cell_sr.border = thin_border
            cell_sr.alignment = left_align
            sr_no += 1

            cell_desc = ws_steel.cell(row=row_num, column=2, value=indent + node["subject"])
            cell_desc.font = bold_font
            cell_desc.border = thin_border
            cell_desc.alignment = left_align
            for c in range(3, steel_total_cols + 1):
                cell = ws_steel.cell(row=row_num, column=c, value="")
                cell.border = thin_border
            row_num += 1
            for r in node["rows"]:
                node_weight += write_leaf_row(r, "", indent, False)

        for child_id in node["children_order"]:
            child = node["children"][child_id]
            node_weight += write_node(child, depth + 1, task_key=task_key)

        if node_weight and depth == 1 and task_key is not None:
            steel_task_totals[task_key] = steel_task_totals.get(task_key, 0) + node_weight

        return node_weight

    for stage_id in node_order:
        stage_node = tree[stage_id]
        stage_weight = write_node(stage_node, 0, task_key=stage_id)
        steel_stage_weights[stage_node["subject"]] = steel_stage_weights.get(stage_node["subject"], 0) + stage_weight
        steel_stage_totals[stage_node["subject"]] = steel_stage_weights[stage_node["subject"]] / 1000
        grand_total_weight += stage_weight

    row_num += 1
    diameter_col_start = 15

    def write_summary_row(label, value_map, bold=False, number_format='0.00'):
        nonlocal row_num
        cell_label = ws_steel.cell(row=row_num, column=2, value=label)
        if bold:
            cell_label.font = bold_font
        cell_label.border = thin_border
        cell_label.alignment = left_align
        for c in [1] + list(range(3, diameter_col_start)):
            cell = ws_steel.cell(row=row_num, column=c, value="")
            cell.border = thin_border
        for idx, d in enumerate(BAR_DIAMETERS):
            val = value_map.get(d, 0)
            cell = ws_steel.cell(row=row_num, column=diameter_col_start + idx, value=val if val else 0)
            if bold:
                cell.font = bold_font
            cell.border = thin_border
            cell.alignment = center_align
            cell.number_format = number_format
        row_num += 1

    weight_kg_by_dia = {d: diameter_length_totals[d] * BAR_WEIGHT_PER_METER[d] for d in BAR_DIAMETERS}
    weight_mt_by_dia = {d: weight_kg_by_dia[d] / 1000 for d in BAR_DIAMETERS}

    write_summary_row("Length In Meter", diameter_length_totals)
    write_summary_row("Weight Per Meter", BAR_WEIGHT_PER_METER, number_format='0.000')
    write_summary_row("Weight In Kg", weight_kg_by_dia)
    write_summary_row("Weight In MT", weight_mt_by_dia, bold=True, number_format='0.000')

    total_mt = sum(weight_mt_by_dia.values())
    cell_label = ws_steel.cell(row=row_num, column=2, value="Total Weight In MT")
    cell_label.font = bold_font
    cell_label.border = thin_border
    total_cell = ws_steel.cell(row=row_num, column=diameter_col_start, value=round(total_mt, 3))
    total_cell.font = bold_font
    total_cell.border = thin_border
    total_cell.number_format = '0.000'
    row_num += 1

    for col in ws_steel.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws_steel.column_dimensions[column].width = (max_length + 2)
    # ============ LEVEL DETAILS SHEET ============
    ws_level = wb.create_sheet(_safe_sheet_name(f"{sheet_prefix} Level Details", used_sheet_names))

    level_headers = [
        "Sr. No", "Task", "Task Subject", "Design", "B.S.", "I.S.", "F.S.", "H.I.", "R.L.",
        "Average R.L.", "Remark"
    ]
    ws_level.append(level_headers)
    for cell in ws_level[1]:
        cell.font = bold_font
        cell.border = thin_border

    row_num = 2
    for row in getattr(doc, "level_details", []):
        if row.task:
            row_type = "header"
        elif row.remark and "average" in row.remark.lower():
            row_type = "average"
        else:
            row_type = "data"

        if row_type == "header":
            ws_level.cell(row=row_num, column=2, value=row.task or "").font = bold_font
            ws_level.cell(row=row_num, column=3, value=row.task_subject or "").font = bold_font
        elif row_type == "average":
            avg_cell = ws_level.cell(row=row_num, column=10, value=flt(row.average_rl))
            avg_cell.font = bold_font
            avg_cell.number_format = '0.000'
            ws_level.cell(row=row_num, column=11, value=row.remark or "").font = bold_font
        else:
            data_fields = [("design", 4), ("bs", 5), ("is", 6), ("fs", 7), ("hi", 8), ("rl", 9)]
            for fieldname, col in data_fields:
                cell = ws_level.cell(row=row_num, column=col, value=flt(row.get(fieldname)))
                cell.number_format = '0.000'
            ws_level.cell(row=row_num, column=11, value=row.remark or "")

        for c in range(1, 12):
            ws_level.cell(row=row_num, column=c).border = thin_border

        row_num += 1

    for col in ws_level.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws_level.column_dimensions[column].width = (max_length + 2)

    # ============ LEVEL DATA SHEET (matrix) ============
    matrix_json = doc.get("level_data_json")
    matrix = None
    if matrix_json:
        try:
            matrix = json.loads(matrix_json)
        except Exception:
            matrix = None

    if matrix and matrix.get("columns"):
        ws_leveldata = wb.create_sheet(_safe_sheet_name(f"{sheet_prefix} Level Data", used_sheet_names))
        columns = matrix["columns"]
        rows = matrix.get("rows", [])

        headers = ["Sr.", "Particular"] + columns
        ws_leveldata.append(headers)
        for cell in ws_leveldata[1]:
            cell.font = bold_font
            cell.border = thin_border

        for idx, row in enumerate(rows, start=1):
            r = idx + 1
            ws_leveldata.cell(row=r, column=1, value=idx).border = thin_border
            pcell = ws_leveldata.cell(row=r, column=2, value=row.get("particular", ""))
            pcell.font = bold_font
            pcell.border = thin_border

            values = row.get("values") or {}
            for c_idx, col in enumerate(columns):
                raw_val = values.get(col, "")
                val = raw_val
                if raw_val not in (None, ""):
                    try:
                        val = flt(raw_val)
                    except Exception:
                        val = raw_val
                cell = ws_leveldata.cell(row=r, column=3 + c_idx, value=val)
                cell.border = thin_border

        for col in ws_leveldata.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws_leveldata.column_dimensions[column].width = (max_length + 2)

    return {
        "project_name": display_project_name,
        "ra_bill": doc.name,
        "abstract_total": abstract_total
    }

@frappe.whitelist()
def export_bulk_ra_excel(bulk_ra_billing):
    bulk_doc = frappe.get_doc("Bulk RA Billing", bulk_ra_billing)

    if not bulk_doc.get("project_details"):
        frappe.throw(_("No project rows found. Please fetch projects and select RA Bills first."))

    with_tax = bool(bulk_doc.get("with_tax"))

    wb = openpyxl.Workbook()
    wb.remove(wb.active) 

    used_sheet_names = set()

    ws_summary = wb.create_sheet(_safe_sheet_name("Summary", used_sheet_names))

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    if with_tax:
        summary_headers = ["Sr. No", "Description", "RA Bill", "Tax Rate", "Amount"]
        amt_col = 5
        rate_col = 4
    else:
        summary_headers = ["Sr. No", "Description", "RA Bill", "Amount"]
        amt_col = 4
        rate_col = None

    total_cols = len(summary_headers)
    last_col = get_column_letter(total_cols)

    ws_summary.merge_cells(f'A1:{last_col}1')
    ws_summary['A1'] = "Billing Summary"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary['A1'].alignment = center_align

    ws_summary.append([])
    ws_summary.append(summary_headers)
    for cell in ws_summary[3]:
        cell.font = bold_font
        cell.border = thin_border
        cell.alignment = center_align if cell.column != 2 else left_align

    summary_row = 4
    grand_total = 0.0

    for idx, row in enumerate(bulk_doc.project_details, start=1):
        if not row.ra_bill:
            continue
        
        project_display_name = frappe.db.get_value("Project", row.project, "project_name") or row.project
        sheet_prefix = f"{idx}-{project_display_name}"
        result = build_ra_sheets_into_workbook(wb, row.ra_bill, sheet_prefix, used_sheet_names)

        ws_summary.cell(row=summary_row, column=1, value=idx).border = thin_border
        ws_summary.cell(row=summary_row, column=2, value=result["project_name"]).border = thin_border
        ws_summary.cell(row=summary_row, column=3, value=result["ra_bill"]).border = thin_border

        if with_tax:
            ws_summary.cell(row=summary_row, column=rate_col, value="").border = thin_border

        amt_cell = ws_summary.cell(row=summary_row, column=amt_col, value=result["abstract_total"])
        amt_cell.border = thin_border
        amt_cell.number_format = '0.00'

        grand_total += result["abstract_total"]
        summary_row += 1

    # ---- Total row (sum of project amounts, before tax) ----
    cell_label = ws_summary.cell(row=summary_row, column=2, value="Total")
    cell_label.font = bold_font
    cell_label.border = thin_border
    ws_summary.cell(row=summary_row, column=1, value="").border = thin_border
    ws_summary.cell(row=summary_row, column=3, value="").border = thin_border
    if with_tax:
        ws_summary.cell(row=summary_row, column=rate_col, value="").border = thin_border

    total_cell = ws_summary.cell(row=summary_row, column=amt_col, value=grand_total)
    total_cell.font = bold_font
    total_cell.border = thin_border
    total_cell.number_format = '0.00'
    summary_row += 1

    # ---- Tax rows (only if With Tax is checked) ----
    tax_total = 0.0
    if with_tax:
        for tax_row in bulk_doc.get("tax_details", []):
            ws_summary.cell(row=summary_row, column=1, value="").border = thin_border
            ws_summary.cell(row=summary_row, column=2, value=tax_row.tax_category or "").border = thin_border
            ws_summary.cell(row=summary_row, column=3, value="").border = thin_border

            rate_cell = ws_summary.cell(row=summary_row, column=rate_col, value=flt(tax_row.tax_rate))
            rate_cell.border = thin_border
            rate_cell.alignment = center_align
            rate_cell.number_format = '0.00"%"'

            tax_amt_cell = ws_summary.cell(row=summary_row, column=amt_col, value=flt(tax_row.tax_amount))
            tax_amt_cell.border = thin_border
            tax_amt_cell.number_format = '0.00'

            tax_total += flt(tax_row.tax_amount)
            summary_row += 1

    # ---- Grand Total row (Total + all tax amounts) ----
    cell_label = ws_summary.cell(row=summary_row, column=2, value="Grand Total")
    cell_label.font = bold_font
    cell_label.border = thin_border
    ws_summary.cell(row=summary_row, column=1, value="").border = thin_border
    ws_summary.cell(row=summary_row, column=3, value="").border = thin_border
    if with_tax:
        ws_summary.cell(row=summary_row, column=rate_col, value="").border = thin_border

    final_grand_total = flt(bulk_doc.get("grand_total")) or (grand_total + tax_total)
    gt_cell = ws_summary.cell(row=summary_row, column=amt_col, value=final_grand_total)
    gt_cell.font = bold_font
    gt_cell.border = thin_border
    gt_cell.number_format = '0.00'

    # highlight grand total row like the reference bill's blue banner
    for c in range(1, total_cols + 1):
        ws_summary.cell(row=summary_row, column=c).fill = openpyxl.styles.PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )

    for col in ws_summary.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws_summary.column_dimensions[column].width = (max_length + 2)

    # Move Summary sheet to the very front
    wb.move_sheet(ws_summary.title, offset=-len(wb.sheetnames))

    file_data = BytesIO()
    wb.save(file_data)

    frappe.response['filename'] = f"Bulk_RA_Billing_{bulk_doc.name}.xlsx"
    frappe.response['filecontent'] = file_data.getvalue()
    frappe.response['type'] = 'binary'